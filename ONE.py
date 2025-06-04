import json
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import threading
import time
import os
import psutil
import re
import openpyxl
import customtkinter as ctk
from selenium import webdriver
from PIL import Image, ImageTk  # Add PIL for image handling
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from datetime import datetime

"""
AutoMessenger ONE - Unified automation tool for sending messages via Onvio Messenger.
Supports multiple models with customizable Excel structures and messages.
"""

# Configuração do tema do customtkinter
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

# Variáveis globais
cancelar = False
log_file_path = None

# Modelos suportados
MODELOS = {
    "ONE": {
        "colunas": ["Código", "Empresa", "Contato Onvio", "Grupo Onvio", "Colaborador", "Evento", "Prazo"],
        "mensagem_padrao": "ONEmessage"
    },
    "ALL": {
        "colunas": ["Codigo", "EMPRESAS", "CONTATO ONVIO", "GRUPO ONVIO"],
        "mensagem_padrao": "Mensagem Padrão"
    },
    "ProrContrato": {
        "colunas": ["Codigo", "Contato Onvio", "Grupo Onvio", "Nome", "Vencimento"],
        "mensagem_padrao": "Prorrogação Contrato"
    },
    "Cobranca": {
        "colunas": ["Código", "Empresa", "Contato Onvio", "Grupo Onvio", "Valor da Parcela", "Data de Vencimento", "Carta de Aviso"],
        "mensagem_padrao": "Cobranca"
    },
    "ComuniCertificado": {
       "colunas": ["Codigo", "Empresa", "Contato Onvio", "Grupo Onvio", "CNPJ", "Vencimento", "Carta de Aviso"],
        "mensagem_padrao": "Cobranca" 
    }
}

def esperar_carregamento_completo(driver):
    try:
        WebDriverWait(driver, 60).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        atualizar_log("Página completamente carregada.")
        return True
    except Exception as e:
        atualizar_log(f"Erro ao esperar carregamento: {str(e)}", cor="vermelho")
        return False

def focar_barra_mensagem_enviar(driver, mensagem):
    try:
        elemento_alvo = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="preview-root"]/div[2]/div[3]/div[1]/div/div[2]/div[2]/div[1]'))
        )
        if elemento_alvo.get_attribute('data-placeholder') == "Mensagem":
            elemento_alvo.click()
            atualizar_log("Barra de Mensagem encontrada e clicada!")
            paragrafos = re.split(r'\n+', mensagem.strip())
            for i, paragrafo in enumerate(paragrafos):
                if i > 0:
                    ActionChains(driver).key_down(Keys.SHIFT).send_keys(Keys.ENTER).key_up(Keys.SHIFT).perform()
                    time.sleep(0.5)
                if cancelar:
                    atualizar_log("Processamento cancelado!", cor="azul")
                    return False
                ActionChains(driver).send_keys(paragrafo).perform()
                time.sleep(0.5)
            atualizar_log("Mensagem formatada inserida com sucesso.")
            
            # Clicar no botão de enviar
            try:
                botao_enviar = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="preview-root"]/div[2]/div[3]/div[3]/div[1]/button'))
                )
                botao_enviar.click()
                atualizar_log("Botão de enviar clicado com sucesso.")
            except:
                atualizar_log("Erro ao clicar no botão de enviar.", cor="vermelho")
                return False
                
            # Clicar no botão de desconsiderar
            try:
                botao_desconsiderar = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="ChatHeader"]/div[2]/div[1]/div[3]/div[1]/button/div'))
                )
                botao_desconsiderar.click()
                atualizar_log("Botão de DESCONSIDERAR clicado com sucesso.")
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, '/html/body/div[4]'))
                )
                desconsiderar = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div/div/div[3]/button[2]'))
                )
                desconsiderar.click()
                time.sleep(2)
                atualizar_log("Mensagem Desconsiderada com Sucesso!", cor="azul")
            except:
                atualizar_log("Erro ao desconsiderar mensagem.", cor="vermelho")
                return False
                
            return True
        atualizar_log("Barra de mensagem não encontrada.")
        return False
    except Exception as e:
        atualizar_log(f"Erro ao focar na barra de mensagem: {str(e)}")
        return False

def encontrar_e_clicar_barra_contatos(driver, contato, grupo):
    try:
        if not esperar_carregamento_completo(driver):
            return False
        if grupo.upper() != "NONE":
            focar_pagina(driver, aba="grupo")
            if focar_barra_endereco_e_navegar(driver, grupo):
                atualizar_log("Navegação aba grupo.")
                return processar_resultados_busca(driver)
        elif contato.upper() != "NONE":
            focar_pagina(driver, aba="contato")
            if focar_barra_endereco_e_navegar(driver, contato):
                atualizar_log("Navegação aba contato.")
                return processar_resultados_busca(driver)
        atualizar_log("Falha na navegação ou busca.", cor="vermelho")
        return False
    except Exception as e:
        atualizar_log(f"Erro ao interagir com a página: {str(e)}", cor="vermelho")
        return False

def enviar_mensagem(driver, contato, grupo, mensagem, codigo, identificador):
    if encontrar_e_clicar_barra_contatos(driver, contato, grupo):
        time.sleep(6)
        if focar_barra_mensagem_enviar(driver, mensagem):
            atualizar_log(f"\nAviso enviado para {contato or grupo}, {codigo} - {identificador}.\n", cor="verde")
            focar_pagina_geral(driver)
            return True
        else:
            atualizar_log(f"Falha ao enviar mensagem para {contato or grupo}. Tentando aba alternativa.", cor="vermelho")
            # Tentar aba alternativa
            if grupo.upper() != "NONE" and contato.upper() != "NONE":
                focar_pagina(driver, aba="contato")
                if focar_barra_endereco_e_navegar(driver, contato):
                    if processar_resultados_busca(driver):
                        atualizar_log("Contato encontrado na aba Contatos.", cor="azul")
                        time.sleep(6)
                        if focar_barra_mensagem_enviar(driver, mensagem):
                            atualizar_log(f"\nAviso enviado para {contato}, {codigo} - {identificador}.\n", cor="verde")
                            focar_pagina_geral(driver)
                            return True
            elif contato.upper() != "NONE" and grupo.upper() != "NONE":
                focar_pagina(driver, aba="grupo")
                if focar_barra_endereco_e_navegar(driver, grupo):
                    if processar_resultados_busca(driver):
                        atualizar_log("Grupo encontrado na aba Grupos.", cor="azul")
                        time.sleep(6)
                        if focar_barra_mensagem_enviar(driver, mensagem):
                            atualizar_log(f"\nAviso enviado para {grupo}, {codigo} - {identificador}.\n", cor="verde")
                            focar_pagina_geral(driver)
                            return True
            atualizar_log(f"Falha ao enviar mensagem para {contato or grupo} em ambas as abas.", cor="vermelho")
    return False

# Funções de Navegação e Automação (reutilizadas do main.py e prorcontrato.py)
def abrir_chrome_com_url(url):
    encerrar_processos_chrome()
    user_data_dir = r"C:\PerfisChrome\automacao"  # o mesmo caminho usado no passo 1
    profile_dir = os.path.join(user_data_dir, "Profile 1")
    # Verificar se o diretório do perfil existe
    if not os.path.exists(profile_dir):
        atualizar_log(f"Perfil 'Profile 1' não encontrado em {user_data_dir}.", cor="vermelho")
        atualizar_log("Um novo perfil será criado. Por favor, faça login na página aberta para continuar.", cor="azul")
        
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
    chrome_options.add_argument("--profile-directory=Profile 1")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-translate")
    chrome_options.add_argument("--lang=pt-BR")
    chrome_options.add_argument("--enable-javascript")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    
    service = Service(ChromeDriverManager().install())
    try:
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.set_page_load_timeout(180)
        driver.get(url)
        atualizar_log(f"Chrome aberto com a URL: {url}")
        return driver
    except Exception as e:
        atualizar_log(f"Erro ao abrir o Chrome: {str(e)}")
        return None

def encerrar_processos_chrome():
    for proc in psutil.process_iter(['name', 'cmdline']):
        if proc.info['name'] == 'chrome.exe':
            try:
                cmdline = proc.info['cmdline'] or []
                cmdline_str = ' '.join(cmdline)
                if '--user-data-dir=C:\\PerfisChrome\\automacao' in cmdline_str and '--profile-directory=Profile 1' in cmdline_str:
                    proc.terminate()
                    atualizar_log(f"Processo Chrome de automação (PID: {proc.pid}) encerrado.")
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
    time.sleep(2)

def focar_barra_endereco_e_navegar(driver, termo_busca):
    try:
        time.sleep(1)
        focused_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="page-content"]/div/div[2]/div[3]/div[1]/div/div/div/input'))
        )
        if focused_element.get_attribute('placeholder') in ["Buscar contatos...", "Buscar grupos..."]:
            focused_element.click()
            atualizar_log(f"Verificando {termo_busca}...")
            valor_atual = focused_element.get_attribute('value')
            if termo_busca != valor_atual:
                focused_element.clear()
                focused_element.send_keys(termo_busca)
                atualizar_log(f"Texto '{termo_busca}' inserido na barra de pesquisa.")
                time.sleep(1)
            else:
                atualizar_log(f"Texto '{termo_busca}' já presente na barra de pesquisa.")
            return True
        atualizar_log("Barra de pesquisa não encontrada.")
        return False
    except Exception as e:
        atualizar_log(f"Erro ao focar na barra de endereço ou navegar: {str(e)}")
        return False

def processar_resultados_busca(driver):
    try:
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="page-content"]/div/div[2]/div[3]/div[2]/div/div[1]'))
        )
        elemento_alvo = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="page-content"]/div/div[2]/div[3]/div[2]/div/div[1]'))
        )
        if elemento_alvo:
            elemento_alvo.click()
            atualizar_log("Clicado no elemento alvo.")
            return True
        atualizar_log("Elemento não encontrado.")
        return False
    except Exception as e:
        atualizar_log(f"Erro ao processar resultados da busca: {str(e)}")
        return False

def focar_pagina(driver, aba="contato"):
    try:
        xpath = '//*[@id="react-tabs-0"]' if aba == "contato" else '//*[@id="react-tabs-2"]'
        elemento = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
        elemento.click()
        atualizar_log(f"Clicado na aba {aba.capitalize()}.", cor="azul")
        time.sleep(3)
        return True
    except Exception as e:
        atualizar_log(f"Erro ao focar na aba {aba}: {str(e)}", cor="vermelho")
        return False

def focar_pagina_geral(driver):
    try:
        elemento = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="page-content"]/div/div[2]/div[1]/div/div/div/div[1]/div/div[1]'))
        )
        elemento.click()
        atualizar_log("Clicado no contato geral.")
        driver.refresh()
        time.sleep(5)
        return True
    except Exception as e:
        atualizar_log(f"Erro ao voltar à lista geral: {str(e)}", cor="vermelho")
        return False

# Funções de Dados
def validar_excel(caminho, modelo):
    try:
        wb = openpyxl.load_workbook(caminho)
        sheet = wb.active
        colunas_excel = [cell.value for cell in sheet[1]]
        colunas_esperadas = MODELOS[modelo]["colunas"]
        if colunas_excel != colunas_esperadas:
            messagebox.showerror("Erro", f"O Excel não corresponde ao modelo {modelo}. Esperado: {colunas_esperadas}")
            return False
        return True
    except Exception as e:
        atualizar_log(f"Erro ao validar Excel: {str(e)}", cor="vermelho")
        return False

def ler_dados_excel(caminho_excel, modelo, linha_inicial=2):
    try:
        wb = openpyxl.load_workbook(caminho_excel)
        sheet = wb.active
        dados = {}
        colunas = MODELOS[modelo]["colunas"]
        for row in sheet.iter_rows(min_row=linha_inicial, values_only=True):
            if row and len(row) >= len(colunas):
                codigo = row[0]
                if modelo == "ProrContrato":
                    nome_contato, nome_grupo, pessoas, vencimentos = row[1:5]
                    if codigo in dados:
                        dados[codigo]['detalhes'].append({'pessoas': pessoas, 'vencimentos': vencimentos})
                    else:
                        dados[codigo] = {
                            'nome_contato': nome_contato,
                            'nome_grupo': nome_grupo,
                            'detalhes': [{'pessoas': pessoas, 'vencimentos': vencimentos}]
                        }
                elif modelo == "Cobranca":
                    codigo, nome, nome_contato, nome_grupo, valores, vencimentos, cartas = row[:7]
                    
                    if not isinstance(cartas, (int, float)) or not 1 <= int(cartas) <= 6:
                        atualizar_log(f"Linha ignorada: Carta de aviso inválida ({cartas}) na linha {row[0]}", cor="vermelho")
                        continue
                    # Se o código da empresa já está no dicionário, adiciona as novas informações à lista
                    if codigo in dados:
                        dados[codigo]['detalhes'].append({
                            'valores': valores,
                            'vencimentos': vencimentos
                        })
                    else:
                        # Caso seja a primeira vez que aparece, inicializa a entrada no dicionário
                        dados[codigo] = {
                            'nome': nome,
                            'nome_contato': nome_contato,
                            'nome_grupo': nome_grupo,
                            'detalhes': [{
                                'valores': valores,
                                'vencimentos': vencimentos
                            }],
                            'cartas': cartas
                        }
                
                elif modelo == "ComuniCertificado":
                    codigo, nome, nome_contato, nome_grupo, cnpj, vencimentos, cartas = row[:7]
                    # Caso seja a primeira vez que aparece, inicializa a entrada no dicionário
                    dados[codigo] = {
                        'nome': nome,
                        'nome_contato': nome_contato,
                        'nome_grupo': nome_grupo,
                        'cnpj': cnpj,
                        'vencimentos': vencimentos,
                        'cartas': cartas
                    }
                    
                else:  # Modelo ALL
                    pessoas, nome_contato, nome_grupo = row[1:4]
                    dados[codigo] = {
                        'empresa': pessoas,
                        'nome_contato': nome_contato,
                        'nome_grupo': nome_grupo
                    }
            else:
                atualizar_log(f"Linha ignorada: {row}")
        return dados if dados else None
    except Exception as e:
        atualizar_log(f"Erro ao ler Excel: {str(e)}", cor="vermelho")
        return None

def extrair_dados(dados, modelo):
    codigos, nome_contatos, nome_grupos = [], [], []
    if modelo == "ProrContrato":
        pessoas, vencimentos = [], []
        for cod, info in dados.items():
            codigos.append(cod)
            nome_contatos.append(info['nome_contato'])
            nome_grupos.append(info['nome_grupo'])
            pessoas_lista = [det['pessoas'] for det in info['detalhes']]
            vencimentos_lista = [det['vencimentos'] for det in info['detalhes']]
            pessoas.append(pessoas_lista)
            vencimentos.append(vencimentos_lista)
        return codigos, nome_contatos, nome_grupos, pessoas, vencimentos
    elif modelo == "Cobranca":
        nome, valores, vencimentos, cartas = [], [], [], []
        # Iterar sobre o dicionário, onde a chave é o código da empresa
        for cod, info in dados.items():
            codigos.append(cod)  # A chave é o código da empresa
            nome.append(info['nome'])  # Extrair o nome
            nome_contatos.append(info['nome_contato'])  # Extrair o nome do contato
            nome_grupos.append(info['nome_grupo'])  # Extrair o nome do grupo
            
            # Para valores e vencimentos, precisamos iterar sobre a lista de detalhes
            valor_total = []
            vencimento_total = []
            
            for detalhe in info['detalhes']:
                valor_total.append(detalhe['valores'])
                vencimento_total.append(detalhe['vencimentos'])
            
            valores.append(valor_total)  # Adicionar a lista de valores associados a esse código
            vencimentos.append(vencimento_total)  # Adicionar a lista de vencimentos associados a esse código
            cartas.append(info['cartas'])  
        
        return codigos, nome, nome_contatos, nome_grupos, valores, vencimentos, cartas 
    
    elif modelo == "ComuniCertificado":
        nome, cnpjs, vencimentos, cartas = [], [], [], []
        # Iterar sobre o dicionário, onde a chave é o código da empresa
        for cod, info in dados.items():
            codigos.append(cod)  # A chave é o código da empresa
            nome.append(info['nome'])  # Extrair o nome
            nome_contatos.append(info['nome_contato'])  # Extrair o nome do contato
            nome_grupos.append(info['nome_grupo'])  # Extrair o nome do grupo
        
            cnpjs.append(info['cnpj'])  # Adicionar a lista de cnpjs associados a esse código
            vencimentos.append(info['vencimentos'])  # Adicionar a lista de vencimentos associados a esse código
            cartas.append(info['cartas'])  
        
        return codigos, nome, nome_contatos, nome_grupos, cnpjs, vencimentos, cartas
    
    else:  # Modelo ALL
        empresas = []
        for cod, info in dados.items():
            codigos.append(cod)
            empresas.append(info['empresa'])
            nome_contatos.append(info['nome_contato'])
            nome_grupos.append(info['nome_grupo'])
        return codigos, empresas, nome_contatos, nome_grupos
    
def formatar_cnpj(cnpj):
    # Remover caracteres não numéricos
    cnpj = ''.join(filter(str.isdigit, cnpj))
    
    # Verificar se o CNPJ tem 14 dígitos
    if len(cnpj) != 14:
        raise ValueError("CNPJ deve conter 14 dígitos")
    
    # Formatar o CNPJ no padrão: XX.XXX.XXX/XXXX-XX
    cnpj_formatado = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"

    return cnpj_formatado

# Funções de Mensagem
def carregar_mensagens():
    try:
        with open("mensagens.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {
            "Mensagem Padrão": "Teste Desconsiderando mensagem",
            "Prorrogação Contrato": "Prezado cliente,\nEspero que estejam bem.\n\nGostaríamos de informar que o contrato de experiência das seguintes pessoas está preste a vencer:\n\n{pessoas_vencimentos}\n\nPara darmos prosseguimento aos devidos registros, solicitamos a gentileza de nos confirmar se haverá prorrogação do contrato ou se ele será encerrado nesta data.\n\nCaso não recebamos um retorno, entenderemos que a prorrogação será realizada automaticamente.\n\nAgradecemos sua atenção.\n\nAtenciosamente,\n\nEquipe DP - C&S."
        }

def salvar_mensagens(mensagens):
    with open("mensagens.json", "w", encoding="utf-8") as f:
        json.dump(mensagens, f, ensure_ascii=False, indent=4)

def mensagem_padrao(modelo, pessoas=None, vencimentos=None, valores=None, carta=None, cnpj=None, nome_empresa=None):
    mensagens = carregar_mensagens()
    msg = mensagens.get(mensagem_selecionada.get(), MODELOS[modelo]["mensagem_padrao"])
    if modelo == "ProrContrato" and pessoas and vencimentos:
        pv = "\n".join([f"{p} se encerrará em {v}" for p, v in zip(pessoas, vencimentos)])
        msg = msg.format(pessoas_vencimentos=pv)
    elif modelo == "Cobranca" and valores and vencimentos and nome_empresa and carta is not None:
        # Formatar valores com vírgula como separador decimal
        valores_formatados = [f"{valor:.2f}".replace('.', ',') for valor in valores]
        total_formatado = f"{sum(valores):.2f}".replace('.', ',')
        # Formatar parcelas
        parcelas = "\n".join([f"Valor: R$ {valor} | Vencimento: {venc}" for valor, venc in zip(valores_formatados, vencimentos)])
        # Selecionar a mensagem com base no número da carta
        msg_key = f"Cobranca_{carta}" if f"Cobranca_{carta}" in mensagens else "Cobranca_1"  # Fallback para carta 1
        msg = mensagens.get(msg_key, mensagens.get("Cobranca_1", "Mensagem de cobrança padrão não encontrada."))
        msg = msg.format(nome=nome_empresa, parcelas=parcelas, total=total_formatado)
    
    elif modelo == "ComuniCertificado":
        cnpj_formatado = formatar_cnpj(cnpj)
         # Selecionar a mensagem com base no número da carta
        msg_key = f"Certificado_{carta}" if f"Certificado_{carta}" in mensagens else "Certificado_1"  # Fallback para carta 1
        msg = mensagens.get(msg_key, mensagens.get("Certificado_1", "Mensagem de cobrança padrão não encontrada."))
        msg = msg.format(nome=nome_empresa, cnpj_formatado=cnpj_formatado, datas=vencimentos)
    else:
        msg = mensagens.get(msg_key, MODELOS[modelo]["mensagem_padrao"])
    return msg

# Funções de Interface
def selecionar_excel():
    arquivo = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if arquivo:
        caminho_excel.set(arquivo)
        modelo = modelo_selecionado.get()
        if modelo and not validar_excel(arquivo, modelo):
            caminho_excel.set("")
        else:
            atualizar_log(f"Arquivo Excel selecionado: {arquivo}")

def atualizar_mensagem_padrao(*args):
    modelo = modelo_selecionado.get()
    if modelo:
        mensagem_padrao_key = MODELOS[modelo]["mensagem_padrao"]
        if modelo == "Cobranca":
            mensagem_padrao_key = "Cobranca"
        elif modelo == "ComuniCertificado":
            mensagem_padrao_key = "Certificado"
        mensagem_selecionada.set(mensagem_padrao_key)

def iniciar_processamento():
    global cancelar
    cancelar = False
    excel = caminho_excel.get()
    modelo = modelo_selecionado.get()
    if not excel or not modelo:
        messagebox.showwarning("Atenção", "Selecione um modelo e um arquivo Excel.")
        return
    try:
        linha = int(entrada_linha_inicial.get())
        if linha < 2:
            raise ValueError("Linha inicial deve ser >= 2")
    except ValueError:
        messagebox.showwarning("Atenção", "Linha inicial deve ser um número inteiro >= 2.")
        return
    atualizar_log("Iniciando processamento...", cor="azul")
    botao_iniciar.configure(state="disabled")
    botao_iniciar_chrome.configure(state="disabled")  # Desativar o botão de Chrome
    inicializar_arquivo_log()
    thread = threading.Thread(target=processar_dados, args=(excel, modelo, linha))
    thread.start()

def processar_dados(excel, modelo, linha_inicial):
    url = "https://app.gestta.com.br/attendance/#/chat/contact-list"
    driver = abrir_chrome_com_url(url)
    if not driver:
        atualizar_log("Não foi possível abrir o Chrome. Processamento abortado.", cor="vermelho")
        finalizar_programa()
        return
    # # Verificar se a página está autenticada
    # try:
    #     WebDriverWait(driver, 10).until(
    #         EC.presence_of_element_located((By.XPATH, '//*[@id="trauth-continue-signin-btn"]'))
    #     )
    #     atualizar_log("Página autenticada, iniciando processamento.", cor="verde")
    # except TimeoutException:
    #     atualizar_log("Falha na autenticação: faça login no Onvio Messenger antes de iniciar o processamento.", cor="vermelho")
    #     driver.quit()
    #     finalizar_programa()
    #     return
    
    time.sleep(10)
    dados = ler_dados_excel(excel, modelo, linha_inicial)
    if not dados:
        atualizar_log("Nenhum dado para processar.", cor="vermelho")
        return
    total_linhas = openpyxl.load_workbook(excel).active.max_row - linha_inicial + 1
    if modelo == "ProrContrato":
        codigos, nome_contatos, nome_grupos, pessoas, vencimentos = extrair_dados(dados, modelo)
        total_contatos = len(codigos)
        for i, (cod, contato, grupo, p, v) in enumerate(zip(codigos, nome_contatos, nome_grupos, pessoas, vencimentos)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                return
            linha_atual = linha_inicial + i
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"Linha: {linha_atual}")
            atualizar_log(f"\nProcessando Empresa: {cod}: Contato: {contato}, Grupo: {grupo}\n", cor="azul")
            mensagem = mensagem_padrao(modelo, pessoas=p, vencimentos=v)
            if enviar_mensagem(driver, contato, grupo, mensagem, cod, p[0]):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo}\n")
            time.sleep(5)

    elif modelo == "Cobranca":
        codigos, nomes, nome_contatos, nome_grupos, valores, vencimentos, cartas = extrair_dados(dados, modelo)
        total_contatos = len(codigos)
        for i, (cod, nome_emp, contato, grupo, p, v, carta) in enumerate(zip(codigos, nomes, nome_contatos, nome_grupos, valores, vencimentos, cartas)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                return
            linha_atual = linha_inicial + i
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"Linha: {linha_atual}")
            atualizar_log(f"\nProcessando contato da empresa {cod} - {nome_emp}: Contato: {contato}, Grupo: {grupo}, Aviso nº: {carta}\n", cor="azul")
            mensagem = mensagem_padrao(modelo, valores=p, vencimentos=v, carta=carta, nome_empresa=nome_emp)
            if enviar_mensagem(driver, contato, grupo, mensagem, cod, nome_emp):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo}\n")
            time.sleep(5)
    
    elif modelo == "ComuniCertificado":
        codigos, nomes, nome_contatos, nome_grupos, cnpjs, vencimentos, cartas = extrair_dados(dados, modelo)
        total_contatos = len(codigos)
        for i, (cod, nome_emp, contato, grupo, c, v, carta) in enumerate(zip(codigos, nomes, nome_contatos, nome_grupos, cnpjs, vencimentos, cartas)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                return
            linha_atual = linha_inicial + i
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"Linha: {linha_atual}")
            atualizar_log(f"\nProcessando contato da empresa {cod} - {nome_emp}: Contato: {contato}, Grupo: {grupo}, Aviso nº: {carta}\n", cor="azul")
            mensagem = mensagem_padrao(modelo, vencimentos=v, carta=carta, cnpj=c, nome_empresa=nome_emp)
            if enviar_mensagem(driver, contato, grupo, mensagem, cod, nome_emp):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo}\n")
            time.sleep(5)
            
    else:  # Modelo ALL
        codigos, empresas, nome_contatos, nome_grupos = extrair_dados(dados, modelo)
        total_contatos = len(codigos)
        for i, (cod, emp, contato, grupo) in enumerate(zip(codigos, empresas, nome_contatos, nome_grupos)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                return
            linha_atual = linha_inicial + i
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"\nProcessando {cod} - {emp}: Contato: {contato}, Grupo: {grupo}\n", cor="azul")
            mensagem = mensagem_padrao(modelo)
            if enviar_mensagem(driver, contato, grupo, mensagem, cod, emp):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo}\n")
            time.sleep(5)
    atualizar_progresso(100, "Concluído")
    atualizar_log("Processamento finalizado!", cor="verde")
    finalizar_programa()

def enviar_mensagem(driver, contato, grupo, mensagem, codigo, identificador):
    if encontrar_e_clicar_barra_contatos(driver, contato, grupo):
        time.sleep(6)
        if focar_barra_mensagem_enviar(driver, mensagem):
            if contato.upper() != "NONE":
                atualizar_log(f"\nAviso enviado para {contato}, {codigo} - {identificador}.\n", cor="verde")
            else: 
                atualizar_log(f"\nAviso enviado para {grupo}, {codigo} - {identificador}.\n", cor="verde")
            focar_pagina_geral(driver)
            return True
        else:
            if contato.upper() != "NONE":
                atualizar_log(f"Falha ao enviar mensagem para {contato}", cor="vermelho")
            else: 
                atualizar_log(f"Falha ao enviar mensagem para {grupo}", cor="vermelho")
            
    return False

def cancelar_processamento():
    global cancelar
    cancelar = True
    atualizar_log("Cancelando processamento...", cor="azul")
    botao_fechar.configure(state="normal")

def fechar_programa():
    janela.quit()

def finalizar_programa():
    messagebox.showinfo("Processo Finalizado", "Processamento concluído!")
    botao_fechar.configure(state="normal")
    botao_iniciar.configure(state="normal")
    botao_iniciar_chrome.configure(state="normal")  # Reativar o botão de Chrome

def abrir_log():
    if log_file_path and os.path.exists(log_file_path):
        os.startfile(log_file_path)
    else:
        messagebox.showinfo("Log não disponível", "Não há log para esta sessão.")

def inicializar_arquivo_log():
    global log_file_path
    log_dir = os.path.join(os.path.expanduser('~'), 'Documents', 'AutoMessengerONE_Logs')
    os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file_path = os.path.join(log_dir, f"automessenger_one_log_{timestamp}.txt")
    with open(log_file_path, 'w', encoding='utf-8') as f:
        f.write(f"=== Log AutoMessenger ONE - {timestamp} ===\n\n")
    return log_file_path

def atualizar_log(mensagem, cor=None):
    log_text.configure(state="normal")
    timestamp = datetime.now().strftime("[%H:%M:%S] ")
    if cor == "vermelho":
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "vermelho")
    elif cor == "verde":
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "verde")
    elif cor == "azul":
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "azul")
    else:
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "preto")
    log_text.configure(state="disabled")
    log_text.see("end")
    if log_file_path and os.path.exists(log_file_path):
        with open(log_file_path, 'a', encoding='utf-8') as f:
            f.write(f"{timestamp}{mensagem}\n")

def atualizar_progresso(valor, texto=""):
    progresso.set(valor / 100)
    progresso_texto.configure(text=texto)
    janela.update_idletasks()

def iniciar_chrome_automacao():
    atualizar_log("Iniciando configuração do Chrome de automação...", cor="azul")
    url = "https://onvio.com.br/staff/#/dashboard-core-center"
    driver = abrir_chrome_com_url(url)
    if driver:
        atualizar_log("Chrome de automação aberto com sucesso. Por favor faça o login, entre no messenger e inicie o processamento.", cor="azul")
        # Não fechamos o driver aqui, deixando-o aberto para o usuário fazer login
    else:
        atualizar_log("Falha ao abrir o Chrome de automação.", cor="vermelho")

# Interface Principal
def main():
    global janela, caminho_excel, modelo_selecionado, mensagem_selecionada, botao_iniciar, botao_fechar, log_text, progresso, progresso_texto, entrada_linha_inicial, botao_iniciar_chrome

    janela = ctk.CTk()
    janela.title("AutoMessenger ONE")
    janela.geometry("700x600")
    janela.resizable(True, True)

    # Set the window icon (use .ico for best compatibility on Windows)
    try:
        janela.iconbitmap("logoOne.ico")  # Replace with your .ico file name
    except:
        # Fallback to .png if .ico fails (works on some platforms)
        icon_image = ctk.CTkImage(Image.open("logoOne.png"), size=(32, 32))  # Adjust size as needed
        janela.iconphoto(False, icon_image)
    
    caminho_excel = ctk.StringVar()
    modelo_selecionado = ctk.StringVar()
    mensagem_selecionada = ctk.StringVar()
    progresso = ctk.DoubleVar()

    # Add the logo to the title frame
    frame_titulo = ctk.CTkFrame(janela, fg_color="transparent")
    frame_titulo.pack(fill="x", padx=10, pady=10)
    
    # Load and display the logo image
    try:
        logo_image = ctk.CTkImage(Image.open("logoOne.png"), size=(50, 50))  # Adjust size as needed
        logo_label = ctk.CTkLabel(frame_titulo, image=logo_image, text="")
        logo_label.pack(pady=(5, 0))
    except Exception as e:
        print(f"Error loading logo image: {e}")  # Log error if image fails to load

    titulo = ctk.CTkLabel(frame_titulo, text="AutoMessenger ONE", font=("Roboto", 16, "bold"))
    titulo.pack(pady=(0, 5))
    
    # frame_titulo = ctk.CTkFrame(janela)
    # frame_titulo.pack(fill="x", padx=10, pady=10)
    # titulo = ctk.CTkLabel(frame_titulo, text="AutoMessenger ONE", font=("Roboto", 16, "bold"))
    # titulo.pack(pady=10)

    frame_selecao = ctk.CTkFrame(janela)
    frame_selecao.pack(fill="x", padx=10, pady=5)

    label_modelo = ctk.CTkLabel(frame_selecao, text="Modelo:")
    label_modelo.grid(row=0, column=0, pady=5, padx=5, sticky="w")
    combo_modelo = ctk.CTkComboBox(frame_selecao, values=list(MODELOS.keys()), variable=modelo_selecionado)
    combo_modelo.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
    # modelo_selecionado.trace("w", atualizar_mensagem_padrao)
    modelo_selecionado.trace_add("write", lambda *args: atualizar_mensagem_padrao())

    label_excel = ctk.CTkLabel(frame_selecao, text="Arquivo Excel:")
    label_excel.grid(row=1, column=0, pady=5, padx=5, sticky="w")
    entrada_excel = ctk.CTkEntry(frame_selecao, textvariable=caminho_excel, width=400)
    entrada_excel.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
    botao_excel = ctk.CTkButton(frame_selecao, text="Selecionar Excel", command=selecionar_excel)
    botao_excel.grid(row=1, column=2, padx=5, pady=5)

    label_linha = ctk.CTkLabel(frame_selecao, text="Iniciar da linha:")
    label_linha.grid(row=2, column=0, pady=5, padx=5, sticky="w")
    entrada_linha_inicial = ctk.CTkEntry(frame_selecao, width=100)
    entrada_linha_inicial.grid(row=2, column=1, padx=5, pady=5, sticky="w")
    entrada_linha_inicial.insert(0, "2")

    frame_mensagem = ctk.CTkFrame(janela)
    frame_mensagem.pack(fill="x", padx=10, pady=5)
    label_mensagem = ctk.CTkLabel(frame_mensagem, text="Mensagem:")
    label_mensagem.grid(row=0, column=0, pady=5, padx=5, sticky="w")
    mensagens = carregar_mensagens()
    combo_mensagem = ctk.CTkComboBox(frame_mensagem, values=list(mensagens.keys()), variable=mensagem_selecionada)
    combo_mensagem.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
    mensagem_selecionada.set(list(mensagens.keys())[0])
    
    def abrir_editor_mensagem():
        janela_editor = ctk.CTkToplevel(janela)
        janela_editor.title("Editor de Mensagens")
        janela_editor.geometry("600x400")

        label_nome = ctk.CTkLabel(janela_editor, text="Nome da Mensagem:")
        label_nome.pack(pady=5)
        entrada_nome = ctk.CTkEntry(janela_editor, width=300)
        entrada_nome.pack(pady=5)

        label_texto = ctk.CTkLabel(janela_editor, text="Texto da Mensagem:")
        label_texto.pack(pady=5)
        texto_mensagem = ctk.CTkTextbox(janela_editor, wrap="word", height=200, width=500)
        texto_mensagem.pack(pady=5)

        def salvar_mensagem():
            nome = entrada_nome.get().strip()
            texto = texto_mensagem.get("1.0", "end").strip()
            if nome and texto:
                mensagens = carregar_mensagens()
                if nome in mensagens and not messagebox.askyesno("Confirmação", f"'{nome}' já existe. Sobrescrever?"):
                    return
                mensagens[nome] = texto
                salvar_mensagens(mensagens)
                combo_mensagem.configure(values=list(mensagens.keys()))
                atualizar_log(f"Mensagem '{nome}' salva!", cor="verde")
                janela_editor.destroy()
            else:
                messagebox.showwarning("Atenção", "Nome e texto são obrigatórios.")

        def remover_mensagem():
            nome = entrada_nome.get().strip()
            if nome:
                mensagens = carregar_mensagens()
                if nome in mensagens and messagebox.askyesno("Confirmação", f"Remover '{nome}'?"):
                    del mensagens[nome]
                    salvar_mensagens(mensagens)
                    combo_mensagem.configure(values=list(mensagens.keys()))
                    mensagem_selecionada.set("")
                    atualizar_log(f"Mensagem '{nome}' removida!", cor="verde")
                    janela_editor.destroy()
                elif nome not in mensagens:
                    messagebox.showwarning("Atenção", "Mensagem não encontrada.")
            else:
                messagebox.showwarning("Atenção", "Digite o nome da mensagem a remover.")

        botao_salvar = ctk.CTkButton(janela_editor, text="Salvar Mensagem", command=salvar_mensagem)
        botao_salvar.pack(pady=5)
        botao_remover = ctk.CTkButton(janela_editor, text="Remover Mensagem", command=remover_mensagem)
        botao_remover.pack(pady=5)

    
    botao_editor = ctk.CTkButton(frame_mensagem, text="Adicionar/Editar Mensagem", command=abrir_editor_mensagem)
    botao_editor.grid(row=0, column=2, padx=5, pady=5)
    
    # Novo botão para iniciar Chrome
    botao_iniciar_chrome = ctk.CTkButton(frame_mensagem, text="Iniciar Chrome de Automação", command=iniciar_chrome_automacao)
    botao_iniciar_chrome.grid(row=0, column=3, padx=5, pady=5)

    frame_botoes = ctk.CTkFrame(janela)
    frame_botoes.pack(fill="x", padx=10, pady=5)
    botao_iniciar = ctk.CTkButton(frame_botoes, text="Iniciar Processamento", command=iniciar_processamento, fg_color="#28a745", hover_color="#218838")
    botao_iniciar.pack(side="left", padx=5, pady=10, expand=True, fill="x")
    botao_cancelar = ctk.CTkButton(frame_botoes, text="Cancelar Processamento", command=cancelar_processamento, fg_color="#dc3545", hover_color="#c82333")
    botao_cancelar.pack(side="left", padx=5, pady=10, expand=True, fill="x")
    botao_fechar = ctk.CTkButton(frame_botoes, text="Fechar Programa", command=fechar_programa, state="disabled", fg_color="#6c757d", hover_color="#5a6268")
    botao_fechar.pack(side="left", padx=5, pady=10, expand=True, fill="x")
    botao_abrir_log = ctk.CTkButton(frame_botoes, text="Abrir Log", command=abrir_log, fg_color="#17a2b8", hover_color="#138496")
    botao_abrir_log.pack(side="left", padx=5, pady=10, expand=True, fill="x")

    frame_progresso = ctk.CTkFrame(janela)
    frame_progresso.pack(fill="x", padx=10, pady=5)
    label_progresso = ctk.CTkLabel(frame_progresso, text="Progresso:")
    label_progresso.pack(side="left", padx=5)
    barra_progresso = ctk.CTkProgressBar(frame_progresso, variable=progresso, width=500)
    barra_progresso.pack(side="left", padx=5, fill="x", expand=True)
    barra_progresso.set(0)
    progresso_texto = ctk.CTkLabel(frame_progresso, text="0/0")
    progresso_texto.pack(side="left", padx=5)

    frame_log = ctk.CTkFrame(janela)
    frame_log.pack(pady=10, padx=10, fill="both", expand=True)
    label_log = ctk.CTkLabel(frame_log, text="Log de execução:")
    label_log.pack(anchor="w", padx=5, pady=5)
    log_text = ctk.CTkTextbox(frame_log, wrap="word", height=250, width=650, fg_color="#F5F5F5")
    log_text.pack(fill="both", expand=True, padx=5, pady=5)
    log_text.tag_config("vermelho", foreground="red")
    log_text.tag_config("verde", foreground="green")
    log_text.tag_config("azul", foreground="blue")
    log_text.tag_config("timestamp", foreground="gray")
    log_text.tag_config("preto", foreground="black")

    atualizar_log("Bem-vindo ao AutoMessenger ONE! Selecione um modelo, Excel e clique em 'Iniciar'.", cor="verde")

    frame_rodape = ctk.CTkFrame(janela, fg_color="transparent")
    frame_rodape.pack(fill="x", padx=10, pady=5)
    label_versao = ctk.CTkLabel(frame_rodape, text="v1.0 | Desenvolvido por Hugo L. Almeida - Equipe de TI", text_color="gray")
    label_versao.pack(side="right", padx=5, pady=5)

    janela.mainloop()

if __name__ == '__main__':
    main()