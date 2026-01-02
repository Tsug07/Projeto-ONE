import json
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import threading
import time
import os
import psutil
import re
import openpyxl
import customtkinter as ctk
from selenium import webdriver
from PIL import Image, ImageTk  # Add PIL for image handling
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from datetime import datetime

"""
AutoMessenger ONE - Unified automation tool for sending messages via Onvio Messenger.
Supports multiple models with customizable Excel structures and messages.
"""

# Configuração do tema do customtkinter
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

# Variáveis globais
cancelar = False
log_file_path = None
anexo_habilitado = None  # Variável para checkbox de anexo
caminho_anexo = None  # Variável para caminho do arquivo anexo
agendamento_ativo = None  # Timer do agendamento
contagem_regressiva_ativa = False  # Flag para contagem regressiva
data_hora_agendada = None  # Data/hora do agendamento
perfil_selecionado = None  # Perfil do Chrome (1 ou 2)
driver_agendamento = None  # Driver do Chrome para agendamento
keep_alive_ativo = False  # Flag para keep-alive
KEEP_ALIVE_INTERVALO = 30 * 60 * 1000  # 30 minutos em milissegundos

# Modelos suportados
MODELOS = {
    # "ONE": {
    #     "colunas": ["Código", "Empresa", "Contato Onvio", "Grupo Onvio", "Colaborador", "Evento", "Prazo"],
    #     "mensagem_padrao": "ONEmessage"
    # },
    "ONE": {
        "colunas": ["Código", "Empresa", "Contato Onvio", "Grupo Onvio", "Caminho"],
        "mensagem_padrao": "ONEmessage"
    },
    "ALL": {
        "colunas": ["Codigo", "Empresa", "Contato Onvio", "Grupo Onvio"],
        "mensagem_padrao": "Mensagem Padrão"
    },
    "ALL_info": {
        "colunas": ["Codigo", "Empresa", "Contato Onvio", "Grupo Onvio", "Competencia"],
        "mensagem_padrao": "ALLinfo"
    },
    # "ProrContrato": {
    #     "colunas": ["Codigo", "Contato Onvio", "Grupo Onvio", "Nome", "Vencimento"],
    #     "mensagem_padrao": "Prorrogação Contrato"
    # },
    "Cobranca": {
        "colunas": ["Código", "Empresa", "Contato Onvio", "Grupo Onvio", "Valor da Parcela", "Data de Vencimento", "Carta de Aviso"],
        "mensagem_padrao": "Cobranca"
    },
    "ComuniCertificado": {
       "colunas": ["Codigo", "Empresa", "Contato Onvio", "Grupo Onvio", "CNPJ", "Vencimento", "Carta de Aviso"],
        "mensagem_padrao": "Cobranca" 
    }
}

def esperar_carregamento_completo(driver):
    try:
        WebDriverWait(driver, 60).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        atualizar_log("Página completamente carregada.")
        return True
    except Exception as e:
        atualizar_log(f"Erro ao esperar carregamento: {str(e)}", cor="vermelho")
        return False

def focar_barra_mensagem_enviar(driver, mensagem, modelo=None, caminhos=None):
    try:
        elemento_alvo = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="preview-root"]/div[2]/div[3]/div[1]/div/div[2]/div[2]/div[1]'))
        )
        if elemento_alvo.get_attribute('data-placeholder') == "Mensagem":
            elemento_alvo.click()
            atualizar_log("Barra de Mensagem encontrada e clicada!")
            paragrafos = re.split(r'\n+', mensagem.strip())
            for i, paragrafo in enumerate(paragrafos):
                if i > 0:
                    ActionChains(driver).key_down(Keys.SHIFT).send_keys(Keys.ENTER).key_up(Keys.SHIFT).perform()
                    time.sleep(0.5)
                if cancelar:
                    atualizar_log("Processamento cancelado!", cor="azul")
                    return False
                ActionChains(driver).send_keys(paragrafo).perform()
                time.sleep(0.5)
            atualizar_log("Mensagem formatada inserida com sucesso.")
            
            # Clicar no botão de enviar
            try:
                botao_enviar = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="preview-root"]/div[2]/div[3]/div[3]/div[1]/button'))
                )
                botao_enviar.click()
                atualizar_log("Botão de enviar clicado com sucesso.", cor="azul")
                time.sleep(5)
            except:
                atualizar_log("Erro ao clicar no botão de enviar.", cor="vermelho")
                return False
            if caminhos:  # Enviar anexo para qualquer modelo que tenha caminhos
                try:
                    input_file = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']"))
                    )
                    # Construir caminhos absolutos e verificar existência
                    caminhos_completos = []
                    for caminho in caminhos:
                        if not os.path.isabs(caminho):
                            caminho_base = os.path.join(os.path.expanduser("~"), "Documents", "Relatorios")
                            caminho_completo = os.path.join(caminho_base, caminho)
                        else:
                            caminho_completo = caminho
                        if not os.path.exists(caminho_completo):
                            atualizar_log(f"Arquivo não encontrado: {caminho_completo}", cor="vermelho")
                            continue
                        caminhos_completos.append(caminho_completo)
                        atualizar_log(f"Preparando anexo: {caminho_completo}")
                    
                    if caminhos_completos:
                        # Enviar todos os arquivos de uma vez
                        input_file.send_keys('\n'.join(caminhos_completos))
                        atualizar_log(f"Arquivos anexados com sucesso: {', '.join(caminhos_completos)}", cor="azul")
                        time.sleep(2 * len(caminhos_completos))  # Ajustar tempo conforme número de arquivos
                        
                        # Clicar no botão de enviar arquivos
                        botao_enviar_arquivo = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="preview-root"]/div[3]/div/div[4]/div[2]/div/button'))
                        )
                        botao_enviar_arquivo.click()
                        atualizar_log("Botão de enviar arquivo clicado com sucesso.")
                        atualizar_log("Aguardando upload do arquivo (pode demorar para vídeos)...", cor="azul")
                        time.sleep(15)  # Delay maior para vídeos carregarem
                    else:
                        atualizar_log("Nenhum arquivo válido para anexar.", cor="vermelho")
                        return False
                except Exception as e:
                    atualizar_log(f"Erro ao anexar arquivos: {e}", cor="vermelho")
                    return False
               
                
            time.sleep(3)       
            # Clicar no botão de desconsiderar
            try:
                botao_desconsiderar = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="ChatHeader"]/div[2]/div[1]/div[3]/div[1]/button/div'))
                )
                botao_desconsiderar.click()
                atualizar_log("Botão de DESCONSIDERAR clicado com sucesso.")
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, '/html/body/div[4]'))
                )
                desconsiderar = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div/div/div[3]/button[2]'))
                )
                desconsiderar.click()
                time.sleep(4)
                atualizar_log("Mensagem Desconsiderada com Sucesso!", cor="azul")
            except:
                atualizar_log("Erro ao desconsiderar mensagem.", cor="vermelho")
                
                # Resolvendo bug de desconsiderar mensagem
                # FECHAR ABA DE TRANSFERENCIA DE MENSAGEM
                atualizar_log("Identificando Bug ...")
                try:
                    #Aba de transferencia de mensagem           
                    janela_transferência = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/div'))
                    )
                    
                    if janela_transferência:
                        atualizar_log("Cancelando Transferencia e corrigindo bug ...")
                        cancelar_transf =  WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div/div/div[3]/button[1]'))
                    )
                        cancelar_transf.click()
                        atualizar_log("Transferencia cancelada, bug corrgido!", cor="azul")
                        return True
                except:
                    atualizar_log("Janela de bug não identificada, e bug não solucionado", cor="vermelho")      
                    return False
                return False
                
            return True
        atualizar_log("Barra de mensagem não encontrada.")
        return False
    except Exception as e:
        atualizar_log(f"Erro ao focar na barra de mensagem: {str(e)}")
        return False

def encontrar_e_clicar_barra_contatos(driver, contato, grupo):
    try:
        if not esperar_carregamento_completo(driver):
            return False
        if grupo.upper() != "NONE":
            focar_pagina(driver, aba="grupo")
            if focar_barra_endereco_e_navegar(driver, grupo):
                atualizar_log("Navegação aba grupo.")
                return processar_resultados_busca(driver)
            else:
                raise TimeoutException("Falha ao navegar na aba grupo")
        elif contato.upper() != "NONE":
            focar_pagina(driver, aba="contato")
            if focar_barra_endereco_e_navegar(driver, contato):
                atualizar_log("Navegação aba contato.")
                return processar_resultados_busca(driver)
            else:
                raise TimeoutException("Falha ao navegar na aba contato")
        atualizar_log("Falha na navegação ou busca: contato e grupo são 'NONE'.", cor="vermelho")
        return False
    except TimeoutException as e:
        atualizar_log(f"Timeout ao tentar localizar na aba inicial: {str(e)}", cor="vermelho")
        raise  # Relança a exceção para ser tratada em enviar_mensagem
    except Exception as e:
        atualizar_log(f"Erro ao interagir com a página: {str(e)}", cor="vermelho")
        return False

def enviar_mensagem(driver, contato, grupo, mensagem, codigo, identificador, modelo=None, caminhos=None):
    try:
        if encontrar_e_clicar_barra_contatos(driver, contato, grupo):
            time.sleep(6)
            if focar_barra_mensagem_enviar(driver, mensagem, modelo, caminhos):
                atualizar_log(f"\nAviso enviado para {contato or grupo}, {codigo} - {identificador}.\n", cor="verde")
                focar_pagina_geral(driver)
                return True
        else:
            raise TimeoutException("Falha inicial na busca de contato ou grupo")
    except TimeoutException:
        atualizar_log(f"Falha ao localizar {contato or grupo}. Tentando aba alternativa.", cor="vermelho")
        # Tentar aba alternativa
        try:
            if grupo.upper() != "NONE":
                atualizar_log("Tentando aba Contatos.", cor="azul")
                focar_pagina(driver, aba="contato")
                # if focar_barra_endereco_e_navegar(driver, contato):
                if processar_resultados_busca(driver):
                    atualizar_log("Contato encontrado na aba Contatos.", cor="azul")
                    time.sleep(6)
                    if focar_barra_mensagem_enviar(driver, mensagem, modelo, caminhos):
                        atualizar_log(f"\nAviso enviado para {contato}, {codigo} - {identificador}.\n", cor="verde")
                        focar_pagina_geral(driver)
                        return True
            elif contato.upper() != "NONE":
                atualizar_log("Tentando aba Grupos.", cor="azul")
                focar_pagina(driver, aba="grupo")
                # if focar_barra_endereco_e_navegar(driver, grupo):
                if processar_resultados_busca(driver):
                    atualizar_log("Grupo encontrado na aba Grupos.", cor="azul")
                    time.sleep(6)
                    if focar_barra_mensagem_enviar(driver, mensagem, modelo, caminhos):
                        atualizar_log(f"\nAviso enviado para {grupo}, {codigo} - {identificador}.\n", cor="verde")
                        focar_pagina_geral(driver)
                        return True
            atualizar_log(f"Falha ao enviar mensagem para {contato or grupo} em ambas as abas.", cor="vermelho")
            return False
        except Exception as e:
            atualizar_log(f"Erro ao tentar aba alternativa: {str(e)}", cor="vermelho")
            return False
    except Exception as e:
        atualizar_log(f"Erro geral ao enviar mensagem: {str(e)}", cor="vermelho")
        return False

# Funções de Navegação e Automação (reutilizadas do main.py e prorcontrato.py)
def obter_perfil_chrome():
    """Retorna o número do perfil baseado na seleção do usuário"""
    return perfil_selecionado.get() if perfil_selecionado else "1"

def obter_user_data_dir():
    """Retorna o diretório de dados do Chrome baseado no perfil selecionado.
    Cada perfil usa um diretório SEPARADO para permitir execução simultânea."""
    perfil = obter_perfil_chrome()
    return rf"C:\PerfisChrome\automacao_perfil{perfil}"

def abrir_chrome_com_url(url):
    # Encerra apenas o Chrome do perfil atual (não interfere no outro perfil)
    encerrar_processos_chrome()
    user_data_dir = obter_user_data_dir()
    perfil = obter_perfil_chrome()

    # Criar diretório se não existir
    if not os.path.exists(user_data_dir):
        os.makedirs(user_data_dir, exist_ok=True)
        atualizar_log(f"Diretório do perfil {perfil} criado.", cor="azul")
        atualizar_log("Por favor, faça login na página aberta para continuar.", cor="azul")

    atualizar_log(f"Usando perfil: {perfil} ({user_data_dir})", cor="azul")

    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-translate")
    chrome_options.add_argument("--lang=pt-BR")
    chrome_options.add_argument("--enable-javascript")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    service = Service(ChromeDriverManager().install())
    try:
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.set_page_load_timeout(180)
        driver.get(url)
        atualizar_log(f"Chrome aberto com a URL: {url}")
        return driver
    except Exception as e:
        atualizar_log(f"Erro ao abrir o Chrome: {str(e)}")
        return None

def encerrar_processos_chrome():
    """Encerra apenas os processos Chrome do perfil selecionado"""
    perfil = obter_perfil_chrome()
    encerrou_algum = False
    for proc in psutil.process_iter(['name', 'cmdline']):
        if proc.info['name'] == 'chrome.exe':
            try:
                cmdline = proc.info['cmdline'] or []
                cmdline_str = ' '.join(cmdline)
                # Encerra apenas o Chrome do diretório do perfil selecionado
                # Verifica com barras normais e invertidas
                if f'automacao_perfil{perfil}' in cmdline_str:
                    proc.terminate()
                    atualizar_log(f"Processo Chrome (Perfil {perfil}) encerrado (PID: {proc.pid}).")
                    encerrou_algum = True
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
    if encerrou_algum:
        time.sleep(2)

def focar_barra_endereco_e_navegar(driver, termo_busca):
    try:
        time.sleep(1)
        focused_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="page-content"]/div/div[2]/div[3]/div[1]/div/div/div/input'))
        )
        if focused_element.get_attribute('placeholder') in ["Buscar contatos...", "Buscar grupos..."]:
            focused_element.click()
            atualizar_log(f"Verificando {termo_busca}...")
            valor_atual = focused_element.get_attribute('value')
            if termo_busca != valor_atual:
                focused_element.clear()
                focused_element.send_keys(termo_busca)
                atualizar_log(f"Texto '{termo_busca}' inserido na barra de pesquisa.")
                time.sleep(1)
            else:
                atualizar_log(f"Texto '{termo_busca}' já presente na barra de pesquisa.")
            return True
        atualizar_log("Barra de pesquisa não encontrada.")
        return False
    except Exception as e:
        atualizar_log(f"Erro ao focar na barra de endereço ou navegar: {str(e)}")
        driver.refresh()
        time.sleep(3)
        focar_barra_endereco_e_navegar(driver, termo_busca)
        return False

def processar_resultados_busca(driver):
    try:
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="page-content"]/div/div[2]/div[3]/div[2]/div/div[1]'))
        )
        elemento_alvo = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="page-content"]/div/div[2]/div[3]/div[2]/div/div[1]'))
        )
        if elemento_alvo:
            elemento_alvo.click()
            atualizar_log("Clicado no elemento alvo.")
            return True
        atualizar_log("Elemento não encontrado.")
        return False
    except Exception as e:
        atualizar_log(f"Erro ao processar resultados da busca: {str(e)}")
        return False

def focar_pagina(driver, aba="contato"):
    try:
        xpath = '//*[@id="react-tabs-0"]' if aba == "contato" else '//*[@id="react-tabs-2"]'
        elemento = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
        elemento.click()
        atualizar_log(f"Clicado na aba {aba.capitalize()}.", cor="azul")
        time.sleep(3)
        return True
    except Exception as e:
        atualizar_log(f"Erro ao focar na aba {aba}: {str(e)}", cor="vermelho")
        return False

def focar_pagina_geral(driver):
    try:
        elemento = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="page-content"]/div/div[2]/div[1]/div/div/div/div[1]/div/div[1]'))
        )
        elemento.click()
        atualizar_log("Clicado no contato geral.")
        driver.refresh()
        time.sleep(5)
        return True
    except Exception as e:
        atualizar_log(f"Erro ao voltar à lista geral: {str(e)}", cor="vermelho")
        driver.refresh()
        focar_pagina_geral(driver)
        return False

# Funções de Dados
def validar_excel(caminho, modelo):
    try:
        wb = openpyxl.load_workbook(caminho)
        sheet = wb.active
        colunas_excel = [cell.value for cell in sheet[1]]
        colunas_esperadas = MODELOS[modelo]["colunas"]
        if colunas_excel != colunas_esperadas:
            messagebox.showerror("Erro", f"O Excel não corresponde ao modelo {modelo}. Esperado: {colunas_esperadas}")
            return False
        return True
    except Exception as e:
        atualizar_log(f"Erro ao validar Excel: {str(e)}", cor="vermelho")
        return False

def ler_dados_excel(caminho_excel, modelo, linha_inicial=2):
    try:
        wb = openpyxl.load_workbook(caminho_excel)
        sheet = wb.active
        dados = {}
        colunas = MODELOS[modelo]["colunas"]
        for row in sheet.iter_rows(min_row=linha_inicial, values_only=True):
            if row and len(row) >= len(colunas):
                codigo = row[0]
                if modelo == "ProrContrato":
                    nome_contato, nome_grupo, pessoas, vencimentos = row[1:5]
                    if codigo in dados:
                        dados[codigo]['detalhes'].append({'pessoas': pessoas, 'vencimentos': vencimentos})
                    else:
                        dados[codigo] = {
                            'nome_contato': nome_contato,
                            'nome_grupo': nome_grupo,
                            'detalhes': [{'pessoas': pessoas, 'vencimentos': vencimentos}]
                        }
                elif modelo == "Cobranca":
                    codigo, nome, nome_contato, nome_grupo, valores, vencimentos, cartas = row[:7]
                    
                    if not isinstance(cartas, (int, float)) or not 1 <= int(cartas) <= 7:
                        atualizar_log(f"Linha ignorada: Carta de aviso inválida ({cartas}) na linha {row[0]}", cor="vermelho")
                        continue
                    # Se o código da empresa já está no dicionário, adiciona as novas informações à lista
                    if codigo in dados:
                        dados[codigo]['detalhes'].append({
                            'valores': valores,
                            'vencimentos': vencimentos
                        })
                    else:
                        # Caso seja a primeira vez que aparece, inicializa a entrada no dicionário
                        dados[codigo] = {
                            'nome': nome,
                            'nome_contato': nome_contato,
                            'nome_grupo': nome_grupo,
                            'detalhes': [{
                                'valores': valores,
                                'vencimentos': vencimentos
                            }],
                            'cartas': cartas
                        }
                
                elif modelo == "ComuniCertificado":
                    codigo, nome, nome_contato, nome_grupo, cnpj, vencimentos, cartas = row[:7]
                    # Caso seja a primeira vez que aparece, inicializa a entrada no dicionário
                    dados[codigo] = {
                        'nome': nome,
                        'nome_contato': nome_contato,
                        'nome_grupo': nome_grupo,
                        'cnpj': cnpj,
                        'vencimentos': vencimentos,
                        'cartas': cartas
                    }
                    
                elif modelo == "ONE":
                    empresa, nome_contato, nome_grupo, caminho = row[1:5]
                    # Agrupar por contato ou grupo (se contato for "NONE")
                    chave = nome_contato if nome_contato.upper() != "NONE" else nome_grupo
                    if chave in dados:
                        dados[chave]['empresas'].append({
                            'codigo': codigo,
                            'empresa': empresa,
                            'caminho': caminho
                        })
                    else:
                        dados[chave] = {
                            'nome_contato': nome_contato,
                            'nome_grupo': nome_grupo,
                            'empresas': [{
                                'codigo': codigo,
                                'empresa': empresa,
                                'caminho': caminho
                            }]
                        }
                
                elif modelo == "ALL_info":
                    empresa, nome_contato, nome_grupo, competencia = row[1:5]
                    # Validar e normalizar valores vazios
                    nome_contato = str(nome_contato) if nome_contato is not None else "NONE"
                    nome_grupo = str(nome_grupo) if nome_grupo is not None else "NONE"
                    competencia = str(competencia) if competencia is not None else ""
                    # Agrupar por contato ou grupo (se contato for "NONE")
                    chave = nome_contato if nome_contato.upper() != "NONE" else nome_grupo
                    if chave in dados:
                        dados[chave]['empresas'].append({
                            'codigo': codigo,
                            'empresa': empresa
                        })
                        # Armazenar competência (assumindo que é a mesma para todas as empresas do mesmo contato)
                        if 'competencia' not in dados[chave]:
                            dados[chave]['competencia'] = competencia
                    else:
                        dados[chave] = {
                            'nome_contato': nome_contato,
                            'nome_grupo': nome_grupo,
                            'competencia': competencia,
                            'empresas': [{
                                'codigo': codigo,
                                'empresa': empresa
                            }]
                        }

                else:  # Modelo ALL
                    empresa, nome_contato, nome_grupo = row[1:4]
                    # Validar e normalizar valores vazios
                    nome_contato = str(nome_contato) if nome_contato is not None else "NONE"
                    nome_grupo = str(nome_grupo) if nome_grupo is not None else "NONE"
                    # Agrupar por contato ou grupo (se contato for "NONE")
                    chave = nome_contato if nome_contato.upper() != "NONE" else nome_grupo
                    if chave in dados:
                        dados[chave]['empresas'].append({
                            'codigo': codigo,
                            'empresa': empresa
                        })
                    else:
                        dados[chave] = {
                            'nome_contato': nome_contato,
                            'nome_grupo': nome_grupo,
                            'empresas': [{
                                'codigo': codigo,
                                'empresa': empresa
                            }]
                        }
            else:
                atualizar_log(f"Linha ignorada: {row}")
        return dados if dados else None
    except Exception as e:
        atualizar_log(f"Erro ao ler Excel: {str(e)}", cor="vermelho")
        return None

def extrair_dados(dados, modelo):
    codigos, nome_contatos, nome_grupos = [], [], []
    # if modelo == "ProrContrato":
    #     pessoas, vencimentos = [], []
    #     for cod, info in dados.items():
    #         codigos.append(cod)
    #         nome_contatos.append(info['nome_contato'])
    #         nome_grupos.append(info['nome_grupo'])
    #         pessoas_lista = [det['pessoas'] for det in info['detalhes']]
    #         vencimentos_lista = [det['vencimentos'] for det in info['detalhes']]
    #         pessoas.append(pessoas_lista)
    #         vencimentos.append(vencimentos_lista)
    #     return codigos, nome_contatos, nome_grupos, pessoas, vencimentos
    if modelo == "Cobranca":
        nome, valores, vencimentos, cartas = [], [], [], []
        # Iterar sobre o dicionário, onde a chave é o código da empresa
        for cod, info in dados.items():
            codigos.append(cod)  # A chave é o código da empresa
            nome.append(info['nome'])  # Extrair o nome
            nome_contatos.append(info['nome_contato'])  # Extrair o nome do contato
            nome_grupos.append(info['nome_grupo'])  # Extrair o nome do grupo
            
            # Para valores e vencimentos, precisamos iterar sobre a lista de detalhes
            valor_total = []
            vencimento_total = []
            
            for detalhe in info['detalhes']:
                valor_total.append(detalhe['valores'])
                vencimento_total.append(detalhe['vencimentos'])
            
            valores.append(valor_total)  # Adicionar a lista de valores associados a esse código
            vencimentos.append(vencimento_total)  # Adicionar a lista de vencimentos associados a esse código
            cartas.append(info['cartas'])  
        
        return codigos, nome, nome_contatos, nome_grupos, valores, vencimentos, cartas 
    
    elif modelo == "ComuniCertificado":
        nome, cnpjs, vencimentos, cartas = [], [], [], []
        # Iterar sobre o dicionário, onde a chave é o código da empresa
        for cod, info in dados.items():
            codigos.append(cod)  # A chave é o código da empresa
            nome.append(info['nome'])  # Extrair o nome
            nome_contatos.append(info['nome_contato'])  # Extrair o nome do contato
            nome_grupos.append(info['nome_grupo'])  # Extrair o nome do grupo
        
            cnpjs.append(info['cnpj'])  # Adicionar a lista de cnpjs associados a esse código
            vencimentos.append(info['vencimentos'])  # Adicionar a lista de vencimentos associados a esse código
            cartas.append(info['cartas'])  
        
        return codigos, nome, nome_contatos, nome_grupos, cnpjs, vencimentos, cartas
    
    elif modelo == "ONE":
        contatos, nome_contatos, nome_grupos, empresas_lista, caminhos_lista = [], [], [], [], []
        for chave, info in dados.items():
            contatos.append(chave)
            nome_contatos.append(info['nome_contato'])
            nome_grupos.append(info['nome_grupo'])
            empresas = [(emp['codigo'], emp['empresa'], emp['caminho']) for emp in info['empresas']]
            empresas_lista.append(empresas)
            caminhos_lista.append([emp['caminho'] for emp in info['empresas']])
        return contatos, nome_contatos, nome_grupos, empresas_lista, caminhos_lista

    elif modelo == "ALL_info":
        contatos, nome_contatos, nome_grupos, empresas_lista, competencias = [], [], [], [], []
        for chave, info in dados.items():
            contatos.append(chave)
            nome_contatos.append(info['nome_contato'])
            nome_grupos.append(info['nome_grupo'])
            competencias.append(info.get('competencia', ''))
            empresas = [(emp['codigo'], emp['empresa']) for emp in info['empresas']]
            empresas_lista.append(empresas)
        return contatos, nome_contatos, nome_grupos, empresas_lista, competencias

    else:  # Modelo ALL
        contatos, nome_contatos, nome_grupos, empresas_lista = [], [], [], []
        for chave, info in dados.items():
            contatos.append(chave)
            nome_contatos.append(info['nome_contato'])
            nome_grupos.append(info['nome_grupo'])
            empresas = [(emp['codigo'], emp['empresa']) for emp in info['empresas']]
            empresas_lista.append(empresas)
        return contatos, nome_contatos, nome_grupos, empresas_lista
    
def formatar_cnpj(cnpj):
    # Remover caracteres não numéricos
    cnpj = ''.join(filter(str.isdigit, cnpj))
    
    # Verificar se o CNPJ tem 14 dígitos
    if len(cnpj) != 14:
        raise ValueError("CNPJ deve conter 14 dígitos")
    
    # Formatar o CNPJ no padrão: XX.XXX.XXX/XXXX-XX
    cnpj_formatado = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"

    return cnpj_formatado

# Funções de Mensagem
def carregar_mensagens():
    try:
        with open("mensagens.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {
            "Mensagem Padrão": "Teste Desconsiderando mensagem",
            "Prorrogação Contrato": "Prezado cliente,\nEspero que estejam bem.\n\nGostaríamos de informar que o contrato de experiência das seguintes pessoas está preste a vencer:\n\n{pessoas_vencimentos}\n\nPara darmos prosseguimento aos devidos registros, solicitamos a gentileza de nos confirmar se haverá prorrogação do contrato ou se ele será encerrado nesta data.\n\nCaso não recebamos um retorno, entenderemos que a prorrogação será realizada automaticamente.\n\nAgradecemos sua atenção.\n\nAtenciosamente,\n\nEquipe DP - C&S."
        }

def salvar_mensagens(mensagens):
    with open("mensagens.json", "w", encoding="utf-8") as f:
        json.dump(mensagens, f, ensure_ascii=False, indent=4)

def mensagem_padrao(modelo, pessoas=None, vencimentos=None, valores=None, carta=None, cnpj=None, nome_empresa=None, competencia=None):
    mensagens = carregar_mensagens()
    msg = mensagens.get(mensagem_selecionada.get(), MODELOS[modelo]["mensagem_padrao"])
    
    # if modelo == "ProrContrato" and pessoas and vencimentos:
    #     pv = "\n".join([f"{p} se encerrará em {v}" for p, v in zip(pessoas, vencimentos)])
    #     msg = msg.format(pessoas_vencimentos=pv)
    if modelo == "Cobranca" and valores and vencimentos and nome_empresa and carta is not None:
        # Formatar valores com vírgula como separador decimal
        valores_formatados = [f"{valor:.2f}".replace('.', ',') for valor in valores]
        total_formatado = f"{sum(valores):.2f}".replace('.', ',')
        # Formatar parcelas
        parcelas = "\n".join([f"Valor: R$ {valor} | Vencimento: {venc}" for valor, venc in zip(valores_formatados, vencimentos)])
        # Selecionar a mensagem com base no número da carta
        msg_key = f"Cobranca_{carta}" if f"Cobranca_{carta}" in mensagens else "Cobranca_1"  # Fallback para carta 1
        msg = mensagens.get(msg_key, mensagens.get("Cobranca_1", "Mensagem de cobrança padrão não encontrada."))
        msg = msg.format(nome=nome_empresa, parcelas=parcelas, total=total_formatado)
    
    elif modelo == "ComuniCertificado":
        cnpj_formatado = formatar_cnpj(cnpj)
         # Selecionar a mensagem com base no número da carta
        msg_key = f"Certificado_{carta}" if f"Certificado_{carta}" in mensagens else "Certificado_1"  # Fallback para carta 1
        msg = mensagens.get(msg_key, mensagens.get("Certificado_1", "Mensagem de cobrança padrão não encontrada."))
        msg = msg.format(nome=nome_empresa, cnpj_formatado=cnpj_formatado, datas=vencimentos)
    
    elif modelo in ["ONE", "ALL", "ALL_info"]:
        # normaliza nome_empresa para lista de nomes
        if isinstance(nome_empresa, list):
            nomes_empresas = nome_empresa
        elif nome_empresa is None:
            nomes_empresas = []
        else:
            nomes_empresas = [nome_empresa]

        # Pegar a mensagem selecionada pelo usuário
        msg_selecionada = mensagem_selecionada.get()

        # Verificar se é uma mensagem que NÃO usa dados (sem placeholders)
        if msg_selecionada == "ONEmessage":
            # Mensagem simples sem dados dinâmicos
            msg = mensagens.get(msg_selecionada, "Mensagem padrão não encontrada.")
        else:
            # Mensagem com dados (Parabens_Regularizado, ALLinfo, etc.)
            if len(nomes_empresas) > 1:
                # Múltiplas empresas - usa versão _multi
                msg_key = f"{msg_selecionada}_multi" if f"{msg_selecionada}_multi" in mensagens else msg_selecionada
                msg = mensagens.get(msg_key, mensagens.get(msg_selecionada, "Mensagem padrão não encontrada."))
                lista_empresas = "\n".join([f". {emp}" for emp in nomes_empresas])
                # Tentar formatar com lista_empresas e competência, se falhar, enviar sem formatação
                try:
                    if competencia:
                        msg = msg.format(empresas=lista_empresas, competencia=competencia)
                    else:
                        msg = msg.format(empresas=lista_empresas)
                except KeyError:
                    pass
            else:
                # Uma única empresa
                msg = mensagens.get(msg_selecionada, "Mensagem padrão não encontrada.")
                nome_unico = nomes_empresas[0] if nomes_empresas else ""
                # Tentar formatar com nome e competência, se falhar, enviar sem formatação
                try:
                    if competencia:
                        msg = msg.format(nome=nome_unico, competencia=competencia)
                    else:
                        msg = msg.format(nome=nome_unico)
                except KeyError:
                    pass
    return msg

# Funções de Interface
def selecionar_excel():
    arquivo = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if arquivo:
        caminho_excel.set(arquivo)
        modelo = modelo_selecionado.get()
        if modelo and not validar_excel(arquivo, modelo):
            caminho_excel.set("")
        else:
            atualizar_log(f"Arquivo Excel selecionado: {arquivo}")

def atualizar_mensagem_padrao(*args):
    modelo = modelo_selecionado.get()
    if modelo:
        mensagem_padrao_key = MODELOS[modelo]["mensagem_padrao"]
        if modelo == "Cobranca":
            mensagem_padrao_key = "Cobranca"
        elif modelo == "ComuniCertificado":
            mensagem_padrao_key = "Certificado"
        mensagem_selecionada.set(mensagem_padrao_key)

def iniciar_processamento():
    global cancelar, driver_agendamento, keep_alive_ativo
    cancelar = False

    # Verificar se há agendamento ativo - não permitir iniciar manualmente
    if agendamento_ativo or keep_alive_ativo:
        messagebox.showwarning("Atenção", "Há um agendamento ativo. Cancele o agendamento antes de iniciar manualmente.")
        return

    excel = caminho_excel.get()
    modelo = modelo_selecionado.get()
    if not excel or not modelo:
        messagebox.showwarning("Atenção", "Selecione um modelo e um arquivo Excel.")
        return
    try:
        linha = int(entrada_linha_inicial.get())
        if linha < 2:
            raise ValueError("Linha inicial deve ser >= 2")
    except ValueError:
        messagebox.showwarning("Atenção", "Linha inicial deve ser um número inteiro >= 2.")
        return
    atualizar_log("Iniciando processamento...", cor="azul")
    botao_iniciar.configure(state="disabled")
    botao_iniciar_chrome.configure(state="disabled")  # Desativar o botão de Chrome
    botao_agendar.configure(state="disabled")  # Desativar agendamento durante processamento
    inicializar_arquivo_log(modelo)
    thread = threading.Thread(target=processar_dados, args=(excel, modelo, linha))
    thread.start()

def processar_dados(excel, modelo, linha_inicial):
    url = "https://app.gestta.com.br/attendance/#/chat/contact-list"
    driver = abrir_chrome_com_url(url)
    if not driver:
        atualizar_log("Não foi possível abrir o Chrome. Processamento abortado.", cor="vermelho")
        finalizar_programa()
        return
    
    time.sleep(10)
    dados = ler_dados_excel(excel, modelo, linha_inicial)
    if not dados:
        atualizar_log("Nenhum dado para processar.", cor="vermelho")
        return
    total_linhas = openpyxl.load_workbook(excel).active.max_row - linha_inicial + 1
    # if modelo == "ProrContrato":
    #     codigos, nome_contatos, nome_grupos, pessoas, vencimentos = extrair_dados(dados, modelo)
    #     total_contatos = len(codigos)
    #     for i, (cod, contato, grupo, p, v) in enumerate(zip(codigos, nome_contatos, nome_grupos, pessoas, vencimentos)):
    #         if cancelar:
    #             atualizar_log("Processamento cancelado!", cor="azul")
    #             return
    #         linha_atual = linha_inicial + i
    #         porcentagem = ((i + 1) / total_contatos) * 100
    #         atualizar_progresso(porcentagem, f"{linha_atual}/{total_linhas + linha_inicial - 1}")
    #         atualizar_log(f"Linha: {linha_atual}")
    #         atualizar_log(f"\nProcessando Empresa: {cod}: Contato: {contato}, Grupo: {grupo}\n", cor="azul")
    #         mensagem = mensagem_padrao(modelo, pessoas=p, vencimentos=v)
    #         if enviar_mensagem(driver, contato, grupo, mensagem, cod, p[0]):
    #             with open(log_file_path, 'a', encoding='utf-8') as f:
    #                 f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo}\n")
    #         time.sleep(5)

    if modelo == "Cobranca":
        codigos, nomes, nome_contatos, nome_grupos, valores, vencimentos, cartas = extrair_dados(dados, modelo)
        total_contatos = len(codigos)
        for i, (cod, nome_emp, contato, grupo, p, v, carta) in enumerate(zip(codigos, nomes, nome_contatos, nome_grupos, valores, vencimentos, cartas)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                return
            linha_atual = linha_inicial + i
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"Linha: {linha_atual}")
            atualizar_log(f"\nProcessando contato da empresa {cod} - {nome_emp}: Contato: {contato}, Grupo: {grupo}, Aviso nº: {carta}\n", cor="azul")
            mensagem = mensagem_padrao(modelo, valores=p, vencimentos=v, carta=carta, nome_empresa=nome_emp)
            if enviar_mensagem(driver, contato, grupo, mensagem, cod, nome_emp):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo}\n")
            time.sleep(5)
    
    elif modelo == "ComuniCertificado":
        codigos, nomes, nome_contatos, nome_grupos, cnpjs, vencimentos, cartas = extrair_dados(dados, modelo)
        total_contatos = len(codigos)
        for i, (cod, nome_emp, contato, grupo, c, v, carta) in enumerate(zip(codigos, nomes, nome_contatos, nome_grupos, cnpjs, vencimentos, cartas)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                return
            linha_atual = linha_inicial + i
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"Linha: {linha_atual}")
            atualizar_log(f"\nProcessando contato da empresa {cod} - {nome_emp}: Contato: {contato}, Grupo: {grupo}, Aviso nº: {carta}\n", cor="azul")
            mensagem = mensagem_padrao(modelo, vencimentos=v, carta=carta, cnpj=c, nome_empresa=nome_emp)
            if enviar_mensagem(driver, contato, grupo, mensagem, cod, nome_emp):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo}\n")
            time.sleep(5)
    
    elif modelo == "ONE":
        contatos, nome_contatos, nome_grupos, empresas_lista, caminhos_lista = extrair_dados(dados, modelo)
        total_contatos = len(contatos)
        linha_atual = linha_inicial
        for i, (contato_key, contato, grupo, empresas, caminhos) in enumerate(zip(contatos, nome_contatos, nome_grupos, empresas_lista, caminhos_lista)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                return
            # Incrementar linha_atual com base no número de empresas processadas
            num_empresas = len(empresas)
            linha_atual_final = linha_atual + num_empresas - 1
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual_final}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"\nProcessando contato {contato_key}: {num_empresas} empresas\n", cor="azul")
            for cod, emp, _ in empresas:
                atualizar_log(f"Empresa: {cod} - {emp}")
             # Monta lista com os nomes das empresas
            nomes_empresas = [emp for _, emp, _ in empresas]

            # Passa a lista de empresas para a mensagem
            mensagem = mensagem_padrao(modelo, nome_empresa=nomes_empresas)
            
            # Enviar uma única mensagem com todos os arquivos
            identificador = ", ".join(nomes_empresas)
            if enviar_mensagem(driver, contato, grupo, mensagem, contato_key, identificador, modelo, caminhos):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo} com {num_empresas} arquivos\n")

            time.sleep(5)
            linha_atual += num_empresas
            
    elif modelo == "ALL_info":
        contatos, nome_contatos, nome_grupos, empresas_lista, competencias = extrair_dados(dados, modelo)
        total_contatos = len(contatos)
        linha_atual = linha_inicial
        for i, (contato_key, contato, grupo, empresas, competencia) in enumerate(zip(contatos, nome_contatos, nome_grupos, empresas_lista, competencias)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                return
            # Incrementar linha_atual com base no número de empresas processadas
            num_empresas = len(empresas)
            linha_atual_final = linha_atual + num_empresas - 1
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual_final}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"\nProcessando contato {contato_key}: {num_empresas} empresas - Competência: {competencia}\n", cor="azul")
            for cod, emp in empresas:
                atualizar_log(f"Empresa: {cod} - {emp}")

            # Monta lista com os nomes das empresas
            nomes_empresas = [emp for _, emp in empresas]

            # Passa a lista de empresas e competência para a mensagem
            mensagem = mensagem_padrao(modelo, nome_empresa=nomes_empresas, competencia=competencia)

            # Enviar uma única mensagem
            identificador = ", ".join(nomes_empresas)
            if enviar_mensagem(driver, contato, grupo, mensagem, contato_key, identificador, modelo):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo} com {num_empresas} empresa(s) - Competência: {competencia}\n")

            time.sleep(5)
            linha_atual += num_empresas

    else:  # Modelo ALL
        contatos, nome_contatos, nome_grupos, empresas_lista = extrair_dados(dados, modelo)
        total_contatos = len(contatos)
        linha_atual = linha_inicial

        # Verificar se há anexo habilitado
        arquivo_anexo = None
        if anexo_habilitado and anexo_habilitado.get() and caminho_anexo and caminho_anexo.get():
            arquivo_anexo = caminho_anexo.get()
            if os.path.exists(arquivo_anexo):
                atualizar_log(f"Anexo configurado: {arquivo_anexo}", cor="azul")
            else:
                atualizar_log(f"Arquivo anexo não encontrado: {arquivo_anexo}", cor="vermelho")
                arquivo_anexo = None

        for i, (contato_key, contato, grupo, empresas) in enumerate(zip(contatos, nome_contatos, nome_grupos, empresas_lista)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                return
            # Incrementar linha_atual com base no número de empresas processadas
            num_empresas = len(empresas)
            linha_atual_final = linha_atual + num_empresas - 1
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual_final}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"\nProcessando contato {contato_key}: {num_empresas} empresas\n", cor="azul")
            for cod, emp in empresas:
                atualizar_log(f"Empresa: {cod} - {emp}")

            # Monta lista com os nomes das empresas
            nomes_empresas = [emp for _, emp in empresas]

            # Passa a lista de empresas para a mensagem
            mensagem = mensagem_padrao(modelo, nome_empresa=nomes_empresas)

            # Enviar uma única mensagem (com anexo opcional)
            identificador = ", ".join(nomes_empresas)
            caminhos_envio = [arquivo_anexo] if arquivo_anexo else None
            if enviar_mensagem(driver, contato, grupo, mensagem, contato_key, identificador, modelo, caminhos_envio):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    anexo_info = " + anexo" if arquivo_anexo else ""
                    f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo} com {num_empresas} empresa(s){anexo_info}\n")

            time.sleep(5)
            linha_atual += num_empresas
    atualizar_progresso(100, "Concluído")
    atualizar_log("Processamento finalizado!", cor="verde")
    finalizar_programa()


def cancelar_processamento():
    global cancelar
    cancelar = True
    atualizar_log("Cancelando processamento...", cor="azul")
    botao_fechar.configure(state="normal")

def fechar_programa():
    global agendamento_ativo, keep_alive_ativo

    # Cancelar agendamento se estiver ativo
    if agendamento_ativo:
        agendamento_ativo.cancel()
        agendamento_ativo = None

    # Parar keep-alive e fechar Chrome do agendamento
    parar_keep_alive()
    fechar_chrome_agendamento()

    janela.quit()

def finalizar_programa():
    messagebox.showinfo("Processo Finalizado", "Processamento concluído!")
    botao_fechar.configure(state="normal")
    botao_iniciar.configure(state="normal")
    botao_iniciar_chrome.configure(state="normal")  # Reativar o botão de Chrome
    botao_agendar.configure(state="normal")  # Reativar agendamento

def finalizar_programa_agendado():
    """Finaliza o programa após processamento agendado e fecha o Chrome"""
    global driver_agendamento
    messagebox.showinfo("Processo Finalizado", "Processamento agendado concluído!")
    botao_fechar.configure(state="normal")
    botao_iniciar.configure(state="normal")
    botao_iniciar_chrome.configure(state="normal")
    botao_agendar.configure(state="normal")

    # Fechar o Chrome do agendamento
    fechar_chrome_agendamento()

def processar_dados_agendado(excel, modelo, linha_inicial):
    """Processa os dados usando o driver já aberto pelo agendamento"""
    global driver_agendamento

    driver = driver_agendamento

    if not driver:
        atualizar_log("Driver não encontrado. Tentando abrir novo Chrome...", cor="vermelho")
        url = "https://app.gestta.com.br/attendance/#/chat/contact-list"
        driver = abrir_chrome_com_url(url)
        if not driver:
            atualizar_log("Não foi possível abrir o Chrome. Processamento abortado.", cor="vermelho")
            finalizar_programa_agendado()
            return

    # Verificar se o driver ainda está ativo
    try:
        driver.current_url
    except:
        atualizar_log("Sessão expirada. Tentando reconectar...", cor="vermelho")
        url = "https://app.gestta.com.br/attendance/#/chat/contact-list"
        driver = abrir_chrome_com_url(url)
        if not driver:
            atualizar_log("Não foi possível reconectar. Processamento abortado.", cor="vermelho")
            finalizar_programa_agendado()
            return
        driver_agendamento = driver

    time.sleep(5)
    dados = ler_dados_excel(excel, modelo, linha_inicial)
    if not dados:
        atualizar_log("Nenhum dado para processar.", cor="vermelho")
        finalizar_programa_agendado()
        return

    total_linhas = openpyxl.load_workbook(excel).active.max_row - linha_inicial + 1
    processamento_cancelado = False

    if modelo == "Cobranca":
        codigos, nomes, nome_contatos, nome_grupos, valores, vencimentos, cartas = extrair_dados(dados, modelo)
        total_contatos = len(codigos)
        for i, (cod, nome_emp, contato, grupo, p, v, carta) in enumerate(zip(codigos, nomes, nome_contatos, nome_grupos, valores, vencimentos, cartas)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                processamento_cancelado = True
                break
            linha_atual = linha_inicial + i
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"Linha: {linha_atual}")
            atualizar_log(f"\nProcessando contato da empresa {cod} - {nome_emp}: Contato: {contato}, Grupo: {grupo}, Aviso nº: {carta}\n", cor="azul")
            mensagem = mensagem_padrao(modelo, valores=p, vencimentos=v, carta=carta, nome_empresa=nome_emp)
            if enviar_mensagem(driver, contato, grupo, mensagem, cod, nome_emp):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo}\n")
            time.sleep(5)

    elif modelo == "ComuniCertificado":
        codigos, nomes, nome_contatos, nome_grupos, cnpjs, vencimentos, cartas = extrair_dados(dados, modelo)
        total_contatos = len(codigos)
        for i, (cod, nome_emp, contato, grupo, c, v, carta) in enumerate(zip(codigos, nomes, nome_contatos, nome_grupos, cnpjs, vencimentos, cartas)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                processamento_cancelado = True
                break
            linha_atual = linha_inicial + i
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"Linha: {linha_atual}")
            atualizar_log(f"\nProcessando contato da empresa {cod} - {nome_emp}: Contato: {contato}, Grupo: {grupo}, Aviso nº: {carta}\n", cor="azul")
            mensagem = mensagem_padrao(modelo, vencimentos=v, carta=carta, cnpj=c, nome_empresa=nome_emp)
            if enviar_mensagem(driver, contato, grupo, mensagem, cod, nome_emp):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo}\n")
            time.sleep(5)

    elif modelo == "ONE":
        contatos, nome_contatos, nome_grupos, empresas_lista, caminhos_lista = extrair_dados(dados, modelo)
        total_contatos = len(contatos)
        linha_atual = linha_inicial
        for i, (contato_key, contato, grupo, empresas, caminhos) in enumerate(zip(contatos, nome_contatos, nome_grupos, empresas_lista, caminhos_lista)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                processamento_cancelado = True
                break
            num_empresas = len(empresas)
            linha_atual_final = linha_atual + num_empresas - 1
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual_final}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"\nProcessando contato {contato_key}: {num_empresas} empresas\n", cor="azul")
            for cod, emp, _ in empresas:
                atualizar_log(f"Empresa: {cod} - {emp}")
            nomes_empresas = [emp for _, emp, _ in empresas]
            mensagem = mensagem_padrao(modelo, nome_empresa=nomes_empresas)
            identificador = ", ".join(nomes_empresas)
            if enviar_mensagem(driver, contato, grupo, mensagem, contato_key, identificador, modelo, caminhos):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo} com {num_empresas} arquivos\n")
            time.sleep(5)
            linha_atual += num_empresas

    elif modelo == "ALL_info":
        contatos, nome_contatos, nome_grupos, empresas_lista, competencias = extrair_dados(dados, modelo)
        total_contatos = len(contatos)
        linha_atual = linha_inicial
        for i, (contato_key, contato, grupo, empresas, competencia) in enumerate(zip(contatos, nome_contatos, nome_grupos, empresas_lista, competencias)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                processamento_cancelado = True
                break
            num_empresas = len(empresas)
            linha_atual_final = linha_atual + num_empresas - 1
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual_final}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"\nProcessando contato {contato_key}: {num_empresas} empresas - Competência: {competencia}\n", cor="azul")
            for cod, emp in empresas:
                atualizar_log(f"Empresa: {cod} - {emp}")
            nomes_empresas = [emp for _, emp in empresas]
            mensagem = mensagem_padrao(modelo, nome_empresa=nomes_empresas, competencia=competencia)
            identificador = ", ".join(nomes_empresas)
            if enviar_mensagem(driver, contato, grupo, mensagem, contato_key, identificador, modelo):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo} com {num_empresas} empresa(s) - Competência: {competencia}\n")
            time.sleep(5)
            linha_atual += num_empresas

    else:  # Modelo ALL
        contatos, nome_contatos, nome_grupos, empresas_lista = extrair_dados(dados, modelo)
        total_contatos = len(contatos)
        linha_atual = linha_inicial

        arquivo_anexo = None
        if anexo_habilitado and anexo_habilitado.get() and caminho_anexo and caminho_anexo.get():
            arquivo_anexo = caminho_anexo.get()
            if os.path.exists(arquivo_anexo):
                atualizar_log(f"Anexo configurado: {arquivo_anexo}", cor="azul")
            else:
                atualizar_log(f"Arquivo anexo não encontrado: {arquivo_anexo}", cor="vermelho")
                arquivo_anexo = None

        for i, (contato_key, contato, grupo, empresas) in enumerate(zip(contatos, nome_contatos, nome_grupos, empresas_lista)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                processamento_cancelado = True
                break
            num_empresas = len(empresas)
            linha_atual_final = linha_atual + num_empresas - 1
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual_final}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"\nProcessando contato {contato_key}: {num_empresas} empresas\n", cor="azul")
            for cod, emp in empresas:
                atualizar_log(f"Empresa: {cod} - {emp}")
            nomes_empresas = [emp for _, emp in empresas]
            mensagem = mensagem_padrao(modelo, nome_empresa=nomes_empresas)
            identificador = ", ".join(nomes_empresas)
            caminhos_envio = [arquivo_anexo] if arquivo_anexo else None
            if enviar_mensagem(driver, contato, grupo, mensagem, contato_key, identificador, modelo, caminhos_envio):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    anexo_info = " + anexo" if arquivo_anexo else ""
                    f.write(f"[{datetime.now()}] ✓ Mensagem enviada para {contato or grupo} com {num_empresas} empresa(s){anexo_info}\n")
            time.sleep(5)
            linha_atual += num_empresas

    if not processamento_cancelado:
        atualizar_progresso(100, "Concluído")
        atualizar_log("Processamento agendado finalizado!", cor="verde")

    # Sempre finalizar e fechar o Chrome, mesmo se cancelado
    finalizar_programa_agendado()

def abrir_log():
    if log_file_path and os.path.exists(log_file_path):
        os.startfile(log_file_path)
    else:
        messagebox.showinfo("Log não disponível", "Não há log para esta sessão.")

def inicializar_arquivo_log(modelo):
    global log_file_path
    # log_dir = os.path.join(os.path.expanduser('~'), 'AutoMessengerONE_Logs')
    log_dir = os.path.join(os.path.dirname(__file__), 'AutoMessengerONE_Logs')
    os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file_path = os.path.join(log_dir, f"{modelo}_log_{timestamp}.txt")
    with open(log_file_path, 'w', encoding='utf-8') as f:
        f.write(f"=== Log AutoMessenger ONE - {timestamp} ===\n\n")
    return log_file_path

def atualizar_log(mensagem, cor=None):
    log_text.configure(state="normal")
    timestamp = datetime.now().strftime("[%H:%M:%S] ")
    if cor == "vermelho":
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "vermelho")
    elif cor == "verde":
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "verde")
    elif cor == "azul":
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "azul")
    else:
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "preto")
    log_text.configure(state="disabled")
    log_text.see("end")
    if log_file_path and os.path.exists(log_file_path):
        with open(log_file_path, 'a', encoding='utf-8') as f:
            f.write(f"{timestamp}{mensagem}\n")

def atualizar_progresso(valor, texto=""):
    progresso.set(valor / 100)
    progresso_texto.configure(text=texto)
    janela.update_idletasks()

def iniciar_chrome_automacao():
    # Verificar se há agendamento ativo
    if agendamento_ativo or keep_alive_ativo:
        messagebox.showwarning("Atenção", "Há um agendamento ativo. Cancele o agendamento antes de abrir o Chrome manualmente.")
        return

    atualizar_log("Iniciando configuração do Chrome de automação...", cor="azul")
    url = "https://onvio.com.br/staff/#/dashboard-core-center"
    driver = abrir_chrome_com_url(url)
    if driver:
        atualizar_log("Chrome de automação aberto com sucesso. Por favor faça o login, entre no messenger e inicie o processamento.", cor="azul")
        # Não fechamos o driver aqui, deixando-o aberto para o usuário fazer login
    else:
        atualizar_log("Falha ao abrir o Chrome de automação.", cor="vermelho")

# Funções de Agendamento
def agendar_processamento():
    global agendamento_ativo, contagem_regressiva_ativa, data_hora_agendada, driver_agendamento

    # Validar campos antes de agendar
    excel = caminho_excel.get()
    modelo = modelo_selecionado.get()
    if not excel or not modelo:
        messagebox.showwarning("Atenção", "Selecione um modelo e um arquivo Excel antes de agendar.")
        return

    try:
        linha = int(entrada_linha_inicial.get())
        if linha < 2:
            raise ValueError("Linha inicial deve ser >= 2")
    except ValueError:
        messagebox.showwarning("Atenção", "Linha inicial deve ser um número inteiro >= 2.")
        return

    # Obter data e hora do agendamento
    try:
        data_str = entrada_data.get().strip()
        hora_str = entrada_hora.get().strip()

        # Validar formato
        if not data_str or not hora_str:
            messagebox.showwarning("Atenção", "Preencha a data e hora do agendamento.")
            return

        # Normalizar data: aceita 02012025 ou 02/01/2025
        data_str = data_str.replace("/", "").replace("-", "").replace(".", "")
        if len(data_str) == 8 and data_str.isdigit():
            data_str = f"{data_str[:2]}/{data_str[2:4]}/{data_str[4:]}"

        # Normalizar hora: aceita 0830 ou 08:30
        hora_str = hora_str.replace(":", "").replace(".", "").replace("-", "")
        if len(hora_str) == 4 and hora_str.isdigit():
            hora_str = f"{hora_str[:2]}:{hora_str[2:]}"

        # Converter para datetime
        data_hora_str = f"{data_str} {hora_str}"
        data_hora_agendada = datetime.strptime(data_hora_str, "%d/%m/%Y %H:%M")

        # Verificar se a data é futura
        agora = datetime.now()
        if data_hora_agendada <= agora:
            messagebox.showwarning("Atenção", "A data/hora deve ser no futuro.")
            return

        # Calcular diferença em segundos
        diferenca = (data_hora_agendada - agora).total_seconds()

        # Cancelar agendamento anterior se existir
        if agendamento_ativo:
            agendamento_ativo.cancel()
            parar_keep_alive()
            fechar_chrome_agendamento()

        # Abrir Chrome e iniciar keep-alive para manter sessão ativa
        atualizar_log("Abrindo Chrome para manter sessão ativa durante o agendamento...", cor="azul")
        driver_agendamento = abrir_chrome_agendamento()

        if not driver_agendamento:
            messagebox.showerror("Erro", "Não foi possível abrir o Chrome. Agendamento cancelado.")
            return

        # Aguardar um pouco para garantir que a página carregou
        time.sleep(5)

        # Iniciar keep-alive (refresh a cada 30 minutos)
        iniciar_keep_alive()

        # Criar novo timer
        agendamento_ativo = threading.Timer(diferenca, executar_agendamento)
        agendamento_ativo.start()

        # Iniciar contagem regressiva
        contagem_regressiva_ativa = True
        atualizar_contagem_regressiva()

        # Log do agendamento
        atualizar_log(f"=" * 50, cor="azul")
        atualizar_log(f"AGENDAMENTO CRIADO COM SUCESSO!", cor="verde")
        atualizar_log(f"Data/Hora programada: {data_hora_agendada.strftime('%d/%m/%Y às %H:%M')}", cor="azul")
        atualizar_log(f"Modelo: {modelo}", cor="azul")
        atualizar_log(f"Excel: {excel}", cor="azul")
        atualizar_log(f"Linha inicial: {linha}", cor="azul")
        atualizar_log(f"Tempo até execução: {formatar_tempo_restante(diferenca)}", cor="azul")
        atualizar_log(f"Keep-alive ativo: Refresh a cada 30 minutos", cor="azul")
        atualizar_log(f"=" * 50, cor="azul")

        # Desabilitar botões
        botao_agendar.configure(state="disabled")
        botao_cancelar_agendamento.configure(state="normal")
        botao_iniciar.configure(state="disabled")
        botao_iniciar_chrome.configure(state="disabled")

        messagebox.showinfo("Agendamento", f"Processamento agendado para:\n{data_hora_agendada.strftime('%d/%m/%Y às %H:%M')}\n\nO Chrome foi aberto e fará refresh automático a cada 30 minutos para manter a sessão ativa.\n\nPor favor, faça login se necessário.")

    except ValueError as e:
        messagebox.showerror("Erro", f"Formato de data/hora inválido.\nUse: DD/MM/AAAA e HH:MM\n\nErro: {str(e)}")

def executar_agendamento():
    global contagem_regressiva_ativa, agendamento_ativo
    contagem_regressiva_ativa = False
    agendamento_ativo = None

    # Parar o keep-alive antes de iniciar o processamento
    parar_keep_alive()

    # Atualizar log
    atualizar_log(f"=" * 50, cor="verde")
    atualizar_log(f"AGENDAMENTO EXECUTANDO!", cor="verde")
    atualizar_log(f"Horário: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}", cor="verde")
    atualizar_log(f"=" * 50, cor="verde")

    # Resetar botões (precisa ser feito na thread principal)
    janela.after(0, lambda: botao_agendar.configure(state="normal"))
    janela.after(0, lambda: botao_cancelar_agendamento.configure(state="disabled"))
    janela.after(0, lambda: label_contagem.configure(text=""))

    # Iniciar processamento usando o driver existente
    janela.after(0, iniciar_processamento_agendado)

def iniciar_processamento_agendado():
    """Inicia o processamento usando o driver já aberto pelo agendamento"""
    global cancelar, driver_agendamento
    cancelar = False
    excel = caminho_excel.get()
    modelo = modelo_selecionado.get()

    if not excel or not modelo:
        messagebox.showwarning("Atenção", "Selecione um modelo e um arquivo Excel.")
        return

    try:
        linha = int(entrada_linha_inicial.get())
        if linha < 2:
            raise ValueError("Linha inicial deve ser >= 2")
    except ValueError:
        messagebox.showwarning("Atenção", "Linha inicial deve ser um número inteiro >= 2.")
        return

    atualizar_log("Iniciando processamento agendado...", cor="azul")
    botao_iniciar.configure(state="disabled")
    botao_iniciar_chrome.configure(state="disabled")
    inicializar_arquivo_log(modelo)

    # Usar o driver existente do agendamento
    thread = threading.Thread(target=processar_dados_agendado, args=(excel, modelo, linha))
    thread.start()

def cancelar_agendamento():
    global agendamento_ativo, contagem_regressiva_ativa, data_hora_agendada

    if agendamento_ativo:
        agendamento_ativo.cancel()
        agendamento_ativo = None

    # Parar keep-alive e fechar Chrome
    parar_keep_alive()
    fechar_chrome_agendamento()

    contagem_regressiva_ativa = False
    data_hora_agendada = None

    # Resetar interface
    botao_agendar.configure(state="normal")
    botao_cancelar_agendamento.configure(state="disabled")
    botao_iniciar.configure(state="normal")
    botao_iniciar_chrome.configure(state="normal")
    label_contagem.configure(text="")

    atualizar_log("Agendamento cancelado pelo usuário.", cor="vermelho")
    messagebox.showinfo("Agendamento", "Agendamento cancelado com sucesso.")

def atualizar_contagem_regressiva():
    global contagem_regressiva_ativa

    if not contagem_regressiva_ativa or not data_hora_agendada:
        return

    agora = datetime.now()
    diferenca = (data_hora_agendada - agora).total_seconds()

    if diferenca <= 0:
        label_contagem.configure(text="Iniciando...")
        return

    # Formatar tempo restante
    texto = formatar_tempo_restante(diferenca)
    label_contagem.configure(text=f"Tempo restante: {texto}")

    # Atualizar a cada segundo
    janela.after(1000, atualizar_contagem_regressiva)

def formatar_tempo_restante(segundos):
    dias = int(segundos // 86400)
    horas = int((segundos % 86400) // 3600)
    minutos = int((segundos % 3600) // 60)
    segs = int(segundos % 60)

    partes = []
    if dias > 0:
        partes.append(f"{dias}d")
    if horas > 0:
        partes.append(f"{horas}h")
    if minutos > 0:
        partes.append(f"{minutos}m")
    partes.append(f"{segs}s")

    return " ".join(partes)

# Funções de Keep-Alive
def iniciar_keep_alive():
    """Inicia o sistema de keep-alive que faz refresh periódico no Chrome"""
    global keep_alive_ativo
    keep_alive_ativo = True
    atualizar_log("Keep-alive iniciado. Refresh a cada 30 minutos.", cor="azul")
    # Agendar primeiro refresh em 30 minutos (não fazer refresh imediato pois o Chrome acabou de abrir)
    janela.after(KEEP_ALIVE_INTERVALO, executar_keep_alive)

def executar_keep_alive():
    """Executa o refresh periódico para manter a sessão ativa"""
    global keep_alive_ativo, driver_agendamento

    if not keep_alive_ativo or not driver_agendamento:
        return

    def fazer_refresh():
        global keep_alive_ativo, driver_agendamento
        try:
            # Verificar se o driver ainda está ativo
            driver_agendamento.current_url

            # Fazer refresh na página
            driver_agendamento.refresh()
            atualizar_log(f"[Keep-alive] Refresh executado às {datetime.now().strftime('%H:%M:%S')}", cor="azul")

            # Agendar próximo refresh (30 minutos) - feito na thread principal
            if keep_alive_ativo:
                janela.after(KEEP_ALIVE_INTERVALO, executar_keep_alive)

        except Exception as e:
            atualizar_log(f"[Keep-alive] Erro no refresh: {str(e)}", cor="vermelho")
            # Tentar reconectar
            try:
                reconectar_chrome_agendamento()
                if keep_alive_ativo:
                    janela.after(KEEP_ALIVE_INTERVALO, executar_keep_alive)
            except:
                atualizar_log("[Keep-alive] Falha ao reconectar. Sessão pode ter expirado.", cor="vermelho")

    # Executar em thread separada para não travar a UI
    thread = threading.Thread(target=fazer_refresh, daemon=True)
    thread.start()

def parar_keep_alive():
    """Para o sistema de keep-alive"""
    global keep_alive_ativo
    keep_alive_ativo = False
    atualizar_log("Keep-alive parado.", cor="azul")

def abrir_chrome_agendamento():
    """Abre o Chrome para o agendamento e retorna o driver"""
    global driver_agendamento

    url = "https://app.gestta.com.br/attendance/#/chat/contact-list"

    # Encerra apenas o Chrome do perfil atual antes de abrir
    encerrar_processos_chrome()

    user_data_dir = obter_user_data_dir()
    perfil = obter_perfil_chrome()

    # Criar diretório se não existir
    if not os.path.exists(user_data_dir):
        os.makedirs(user_data_dir, exist_ok=True)
        atualizar_log(f"Diretório do perfil {perfil} criado.", cor="azul")

    atualizar_log(f"Abrindo Chrome para agendamento (Perfil: {perfil})...", cor="azul")

    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-translate")
    chrome_options.add_argument("--lang=pt-BR")
    chrome_options.add_argument("--enable-javascript")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    service = Service(ChromeDriverManager().install())
    try:
        driver_agendamento = webdriver.Chrome(service=service, options=chrome_options)
        driver_agendamento.set_page_load_timeout(180)
        driver_agendamento.get(url)
        atualizar_log(f"Chrome aberto no Messenger.", cor="verde")
        atualizar_log("Por favor, faça login se necessário.", cor="azul")
        return driver_agendamento
    except Exception as e:
        atualizar_log(f"Erro ao abrir Chrome: {str(e)}", cor="vermelho")
        driver_agendamento = None
        return None

def reconectar_chrome_agendamento():
    """Tenta reconectar o Chrome caso a sessão tenha caído"""
    global driver_agendamento

    atualizar_log("Tentando reconectar Chrome...", cor="azul")

    try:
        if driver_agendamento:
            driver_agendamento.quit()
    except:
        pass

    driver_agendamento = abrir_chrome_agendamento()
    if driver_agendamento:
        atualizar_log("Chrome reconectado com sucesso!", cor="verde")
    else:
        atualizar_log("Falha ao reconectar Chrome.", cor="vermelho")

def fechar_chrome_agendamento():
    """Fecha o Chrome do agendamento"""
    global driver_agendamento, keep_alive_ativo

    keep_alive_ativo = False

    if driver_agendamento:
        try:
            driver_agendamento.quit()
            atualizar_log("Chrome do agendamento fechado.", cor="azul")
        except:
            pass
        driver_agendamento = None

# Interface Principal
def main():
    global janela, caminho_excel, modelo_selecionado, mensagem_selecionada, botao_iniciar, botao_fechar, log_text, progresso, progresso_texto, entrada_linha_inicial, botao_iniciar_chrome, anexo_habilitado, caminho_anexo
    global entrada_data, entrada_hora, botao_agendar, botao_cancelar_agendamento, label_contagem
    global perfil_selecionado

    janela = ctk.CTk()
    janela.title("AutoMessenger ONE")
    janela.geometry("1150x340")
    janela.resizable(True, True)

    # # Set the window icon (use .ico for best compatibility on Windows)
    # try:
    #     janela.iconbitmap("logoOne.ico")  # Replace with your .ico file name
    # except:
    #     # Fallback to .png if .ico fails (works on some platforms)
    #     icon_image = ctk.CTkImage(Image.open("logoOne.png"), size=(32, 32))  # Adjust size as needed
    #     janela.iconphoto(False, icon_image)
    def resource_path(relative_path):
        """Usa caminho relativo compatível com PyInstaller"""
        try:
            base_path = sys._MEIPASS  # PyInstaller usa isso
        except Exception:
            base_path = os.path.abspath(".")

        return os.path.join(base_path, relative_path)

    try:
        janela.iconbitmap(resource_path("logoOne.ico"))
    except:
        try:
            icon_image = ctk.CTkImage(Image.open(resource_path("logoOne.png")), size=(32, 32))
            janela.iconphoto(False, icon_image)
        except Exception as e:
            print(f"Falha ao carregar ícone: {e}")
    
    caminho_excel = ctk.StringVar()
    modelo_selecionado = ctk.StringVar()
    mensagem_selecionada = ctk.StringVar()
    progresso = ctk.DoubleVar()

    # ========== LAYOUT HORIZONTAL: 3 COLUNAS (40% / 25% / 35%) ==========

    # Frame do título (topo)
    frame_titulo = ctk.CTkFrame(janela, fg_color="transparent")
    frame_titulo.pack(fill="x", padx=10, pady=(5, 2))

    # Logo e título lado a lado
    try:
        logo_image = ctk.CTkImage(Image.open(resource_path("logoOne.png")), size=(28, 28))
        logo_label = ctk.CTkLabel(frame_titulo, image=logo_image, text="")
        logo_label.pack(side="left", padx=(5, 3))
    except Exception as e:
        print(f"Error loading logo image: {e}")

    titulo = ctk.CTkLabel(frame_titulo, text="AutoMessenger ONE", font=("Roboto", 14, "bold"))
    titulo.pack(side="left", padx=3)

    # Frame principal com 3 colunas (40% / 25% / 35%)
    frame_principal = ctk.CTkFrame(janela, fg_color="transparent")
    frame_principal.pack(fill="both", expand=True, padx=5, pady=2)
    frame_principal.grid_columnconfigure(0, weight=40)  # Coluna 1: Configurações (40%)
    frame_principal.grid_columnconfigure(1, weight=25)  # Coluna 2: Ações (25%)
    frame_principal.grid_columnconfigure(2, weight=35)  # Coluna 3: Log (35%)
    frame_principal.grid_rowconfigure(0, weight=1)

    # ========== COLUNA 1: Configurações (40%) - Campos agrupados horizontalmente ==========
    frame_col1 = ctk.CTkFrame(frame_principal)
    frame_col1.grid(row=0, column=0, sticky="nsew", padx=3, pady=0)

    # Linha 1: Modelo + Perfil
    frame_row1 = ctk.CTkFrame(frame_col1, fg_color="transparent")
    frame_row1.pack(fill="x", padx=10, pady=(10, 4))

    ctk.CTkLabel(frame_row1, text="Modelo:", font=("Roboto", 11)).pack(side="left")
    combo_modelo = ctk.CTkComboBox(frame_row1, values=list(MODELOS.keys()), variable=modelo_selecionado, width=130)
    combo_modelo.pack(side="left", padx=(5, 15))
    modelo_selecionado.trace_add("write", lambda *args: atualizar_mensagem_padrao())

    ctk.CTkLabel(frame_row1, text="Perfil:", font=("Roboto", 11)).pack(side="left")
    perfil_selecionado = ctk.StringVar(value="1")
    combo_perfil = ctk.CTkComboBox(frame_row1, values=["1", "2"], variable=perfil_selecionado, width=50)
    combo_perfil.pack(side="left", padx=(5, 5))
    botao_iniciar_chrome = ctk.CTkButton(frame_row1, text="Chrome", command=iniciar_chrome_automacao, width=60)
    botao_iniciar_chrome.pack(side="left")

    # Linha 2: Excel + botão
    frame_row2 = ctk.CTkFrame(frame_col1, fg_color="transparent")
    frame_row2.pack(fill="x", padx=10, pady=4)

    ctk.CTkLabel(frame_row2, text="Excel:", font=("Roboto", 11)).pack(side="left")
    entrada_excel = ctk.CTkEntry(frame_row2, textvariable=caminho_excel, width=280)
    entrada_excel.pack(side="left", padx=(5, 3), fill="x", expand=True)
    botao_excel = ctk.CTkButton(frame_row2, text="...", command=selecionar_excel, width=28)
    botao_excel.pack(side="left")

    # Linha 3: Linha + Mensagem + Edit
    frame_row3 = ctk.CTkFrame(frame_col1, fg_color="transparent")
    frame_row3.pack(fill="x", padx=10, pady=4)

    ctk.CTkLabel(frame_row3, text="Linha:", font=("Roboto", 11)).pack(side="left")
    entrada_linha_inicial = ctk.CTkEntry(frame_row3, width=45)
    entrada_linha_inicial.pack(side="left", padx=(5, 15))
    entrada_linha_inicial.insert(0, "2")

    ctk.CTkLabel(frame_row3, text="Mensagem:", font=("Roboto", 11)).pack(side="left")
    mensagens = carregar_mensagens()
    combo_mensagem = ctk.CTkComboBox(frame_row3, values=list(mensagens.keys()), variable=mensagem_selecionada, width=140)
    combo_mensagem.pack(side="left", padx=(5, 3))
    mensagem_selecionada.set(list(mensagens.keys())[0])
    
    def abrir_editor_mensagem():
        janela_editor = ctk.CTkToplevel(janela)
        janela_editor.title("Editor de Mensagens")
        janela_editor.geometry("600x400")

        label_nome = ctk.CTkLabel(janela_editor, text="Nome da Mensagem:")
        label_nome.pack(pady=5)
        entrada_nome = ctk.CTkEntry(janela_editor, width=300)
        entrada_nome.pack(pady=5)

        label_texto = ctk.CTkLabel(janela_editor, text="Texto da Mensagem:")
        label_texto.pack(pady=5)
        texto_mensagem = ctk.CTkTextbox(janela_editor, wrap="word", height=200, width=500)
        texto_mensagem.pack(pady=5)

        def salvar_mensagem():
            nome = entrada_nome.get().strip()
            texto = texto_mensagem.get("1.0", "end").strip()
            if nome and texto:
                mensagens = carregar_mensagens()
                if nome in mensagens and not messagebox.askyesno("Confirmação", f"'{nome}' já existe. Sobrescrever?"):
                    return
                mensagens[nome] = texto
                salvar_mensagens(mensagens)
                combo_mensagem.configure(values=list(mensagens.keys()))
                atualizar_log(f"Mensagem '{nome}' salva!", cor="verde")
                janela_editor.destroy()
            else:
                messagebox.showwarning("Atenção", "Nome e texto são obrigatórios.")

        def remover_mensagem():
            nome = entrada_nome.get().strip()
            if nome:
                mensagens = carregar_mensagens()
                if nome in mensagens and messagebox.askyesno("Confirmação", f"Remover '{nome}'?"):
                    del mensagens[nome]
                    salvar_mensagens(mensagens)
                    combo_mensagem.configure(values=list(mensagens.keys()))
                    mensagem_selecionada.set("")
                    atualizar_log(f"Mensagem '{nome}' removida!", cor="verde")
                    janela_editor.destroy()
                elif nome not in mensagens:
                    messagebox.showwarning("Atenção", "Mensagem não encontrada.")
            else:
                messagebox.showwarning("Atenção", "Digite o nome da mensagem a remover.")

        botao_salvar = ctk.CTkButton(janela_editor, text="Salvar Mensagem", command=salvar_mensagem)
        botao_salvar.pack(pady=5)
        botao_remover = ctk.CTkButton(janela_editor, text="Remover Mensagem", command=remover_mensagem)
        botao_remover.pack(pady=5)

    
    botao_editor = ctk.CTkButton(frame_row3, text="Edit", command=abrir_editor_mensagem, width=28)
    botao_editor.pack(side="left")

    # Linha 4: Anexo + botão
    frame_row4 = ctk.CTkFrame(frame_col1, fg_color="transparent")
    frame_row4.pack(fill="x", padx=10, pady=(4, 10))

    anexo_habilitado = ctk.BooleanVar(value=False)
    caminho_anexo = ctk.StringVar()

    checkbox_anexo = ctk.CTkCheckBox(frame_row4, text="Anexo:", variable=anexo_habilitado, command=lambda: toggle_anexo(), width=20)
    checkbox_anexo.pack(side="left")
    entrada_anexo = ctk.CTkEntry(frame_row4, textvariable=caminho_anexo, width=200, state="disabled")
    entrada_anexo.pack(side="left", padx=(5, 3), fill="x", expand=True)

    def selecionar_anexo():
        arquivo = filedialog.askopenfilename(filetypes=[
            ("Vídeos", "*.mp4 *.avi *.mov *.mkv *.wmv *.webm"),
            ("Imagens", "*.jpg *.jpeg *.png *.gif *.bmp *.webp"),
            ("PDF", "*.pdf"),
            ("Documentos", "*.doc *.docx *.xls *.xlsx *.ppt *.pptx"),
            ("Todos os arquivos", "*.*")
        ])
        if arquivo:
            caminho_anexo.set(arquivo)
            atualizar_log(f"Arquivo anexo selecionado: {arquivo}")

    botao_anexo = ctk.CTkButton(frame_row4, text="...", command=selecionar_anexo, state="disabled", width=28)
    botao_anexo.pack(side="left")

    def toggle_anexo():
        if anexo_habilitado.get():
            entrada_anexo.configure(state="normal")
            botao_anexo.configure(state="normal")
        else:
            entrada_anexo.configure(state="disabled")
            botao_anexo.configure(state="disabled")
            caminho_anexo.set("")

    # ========== COLUNA 2: Ações (25%) - Centralizado com padding 20px ==========
    frame_col2 = ctk.CTkFrame(frame_principal)
    frame_col2.grid(row=0, column=1, sticky="nsew", padx=3, pady=0)

    # Container interno centralizado
    frame_col2_inner = ctk.CTkFrame(frame_col2, fg_color="transparent")
    frame_col2_inner.place(relx=0.5, rely=0.5, anchor="center")

    # Agendamento compacto e centralizado
    label_agendar = ctk.CTkLabel(frame_col2_inner, text="Agendamento", font=("Roboto", 11, "bold"))
    label_agendar.pack(pady=(0, 5))

    frame_agendar_row = ctk.CTkFrame(frame_col2_inner, fg_color="transparent")
    frame_agendar_row.pack(pady=2)
    entrada_data = ctk.CTkEntry(frame_agendar_row, width=75, placeholder_text="DD/MM/AAAA")
    entrada_data.pack(side="left", padx=2)
    entrada_hora = ctk.CTkEntry(frame_agendar_row, width=50, placeholder_text="HH:MM")
    entrada_hora.pack(side="left", padx=2)

    frame_agendar_btns = ctk.CTkFrame(frame_col2_inner, fg_color="transparent")
    frame_agendar_btns.pack(pady=3)
    botao_agendar = ctk.CTkButton(frame_agendar_btns, text="Agendar", command=agendar_processamento, fg_color="#6f42c1", hover_color="#5a32a3", width=70)
    botao_agendar.pack(side="left", padx=2)
    botao_cancelar_agendamento = ctk.CTkButton(frame_agendar_btns, text="Cancelar", command=cancelar_agendamento, fg_color="#fd7e14", hover_color="#e06b0a", width=70, state="disabled")
    botao_cancelar_agendamento.pack(side="left", padx=2)

    label_contagem = ctk.CTkLabel(frame_col2_inner, text="", text_color="#6f42c1", font=("Roboto", 9, "bold"))
    label_contagem.pack(pady=2)

    # Separador
    ctk.CTkLabel(frame_col2_inner, text="─" * 18, text_color="gray").pack(pady=8)

    # Botões de Ação - Grid 2x2 equilibrado
    frame_btns_grid = ctk.CTkFrame(frame_col2_inner, fg_color="transparent")
    frame_btns_grid.pack()

    botao_iniciar = ctk.CTkButton(frame_btns_grid, text="Iniciar", command=iniciar_processamento, fg_color="#28a745", hover_color="#218838", width=70)
    botao_iniciar.grid(row=0, column=0, padx=3, pady=3)
    botao_cancelar = ctk.CTkButton(frame_btns_grid, text="Cancelar", command=cancelar_processamento, fg_color="#dc3545", hover_color="#c82333", width=70)
    botao_cancelar.grid(row=0, column=1, padx=3, pady=3)
    botao_fechar = ctk.CTkButton(frame_btns_grid, text="Fechar", command=fechar_programa, state="disabled", fg_color="#6c757d", hover_color="#5a6268", width=70)
    botao_fechar.grid(row=1, column=0, padx=3, pady=3)
    botao_abrir_log = ctk.CTkButton(frame_btns_grid, text="Log", command=abrir_log, fg_color="#17a2b8", hover_color="#138496", width=70)
    botao_abrir_log.grid(row=1, column=1, padx=3, pady=3)

    # ========== COLUNA 3: Log (35%) - Progresso fixo no topo ==========
    frame_col3 = ctk.CTkFrame(frame_principal)
    frame_col3.grid(row=0, column=2, sticky="nsew", padx=3, pady=0)

    # Progresso fixo no topo
    frame_progresso = ctk.CTkFrame(frame_col3, fg_color="transparent")
    frame_progresso.pack(fill="x", padx=10, pady=(10, 5))
    label_progresso = ctk.CTkLabel(frame_progresso, text="Progresso:", font=("Roboto", 10))
    label_progresso.pack(side="left", padx=(0, 5))
    barra_progresso = ctk.CTkProgressBar(frame_progresso, variable=progresso)
    barra_progresso.pack(side="left", fill="x", expand=True, padx=(0, 5))
    barra_progresso.set(0)
    progresso_texto = ctk.CTkLabel(frame_progresso, text="0/0")
    progresso_texto.pack(side="left")

    # Log ocupando área máxima
    label_log = ctk.CTkLabel(frame_col3, text="Log:", font=("Roboto", 10, "bold"))
    label_log.pack(anchor="w", padx=10, pady=(0, 2))
    log_text = ctk.CTkTextbox(frame_col3, wrap="word", fg_color="#F5F5F5")
    log_text.pack(fill="both", expand=True, padx=10, pady=(0, 10))
    log_text.tag_config("vermelho", foreground="red")
    log_text.tag_config("verde", foreground="green")
    log_text.tag_config("azul", foreground="blue")
    log_text.tag_config("timestamp", foreground="gray")
    log_text.tag_config("preto", foreground="black")

    atualizar_log("Bem-vindo ao AutoMessenger ONE!", cor="verde")

    # Rodapé compacto
    frame_rodape = ctk.CTkFrame(janela, fg_color="transparent")
    frame_rodape.pack(fill="x", padx=5, pady=1)
    label_versao = ctk.CTkLabel(frame_rodape, text="v1.0 | Hugo L. Almeida", text_color="gray", font=("Roboto", 9))
    label_versao.pack(side="right", padx=5)

    janela.mainloop()

if __name__ == '__main__':
    main()